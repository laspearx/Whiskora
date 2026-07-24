from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path
from xml.etree.ElementTree import iterparse
from zipfile import ZipFile
import re

import openpyxl
from openpyxl import Workbook


SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import build_stock_cover_v3_all_months as core  # noqa: E402
import create_stock_cover_v4_historical_simulation as hist  # noqa: E402
import validate_stock_cover_monthly_inputs as validator  # noqa: E402


OUTPUT = core.REPORT_DIR / "outputs" / "validation" / "Historical Edit V4 Impact Summary.xlsx"
NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NS_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
NS_DOC_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def n(value):
    if value is None:
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def product_rsp(label, product_name, product_by_norm, rsp_by_name):
    master = product_by_norm.get(product_name) or {}
    return rsp_by_name.get((label, product_name)) or master.get("marketPrice") or 0


def col_name(ref):
    return re.sub(r"\d+", "", ref or "")


def load_shared_strings(zf):
    try:
        with zf.open("xl/sharedStrings.xml") as handle:
            strings = []
            text_parts = []
            for event, elem in iterparse(handle, events=("start", "end")):
                if event == "start" and elem.tag == f"{NS_MAIN}si":
                    text_parts = []
                elif event == "end" and elem.tag == f"{NS_MAIN}t":
                    text_parts.append(elem.text or "")
                elif event == "end" and elem.tag == f"{NS_MAIN}si":
                    strings.append("".join(text_parts))
                    elem.clear()
            return strings
    except KeyError:
        return []


def sheet_path(zf, sheet_name):
    target_id = None
    with zf.open("xl/workbook.xml") as handle:
        for _event, elem in iterparse(handle, events=("end",)):
            if elem.tag == f"{NS_MAIN}sheet" and elem.attrib.get("name") == sheet_name:
                target_id = elem.attrib.get(f"{NS_DOC_REL}id")
                break
    if not target_id:
        raise ValueError(f"Sheet not found: {sheet_name}")
    with zf.open("xl/_rels/workbook.xml.rels") as handle:
        for _event, elem in iterparse(handle, events=("end",)):
            if elem.tag == f"{NS_REL}Relationship" and elem.attrib.get("Id") == target_id:
                target = elem.attrib["Target"].lstrip("/")
                return target if target.startswith("xl/") else "xl/" + target
    raise ValueError(f"Relationship not found for {sheet_name}")


def cell_value(cell, shared_strings):
    value = cell.find(f"{NS_MAIN}v")
    if value is None:
        inline = cell.find(f"{NS_MAIN}is/{NS_MAIN}t")
        return inline.text if inline is not None else None
    text = value.text
    if cell.attrib.get("t") == "s":
        return shared_strings[int(text)]
    return text


def iter_xlsx_rows(path, sheet_name="Campaign 299"):
    with ZipFile(path) as zf:
        shared_strings = load_shared_strings(zf)
        raw_sheet = sheet_path(zf, sheet_name)
        headers = None
        with zf.open(raw_sheet) as handle:
            for _event, row in iterparse(handle, events=("end",)):
                if row.tag != f"{NS_MAIN}row":
                    continue
                values = {}
                for cell in row:
                    values[col_name(cell.attrib.get("r"))] = cell_value(cell, shared_strings)
                if headers is None:
                    headers = values
                    row.clear()
                    continue
                yield {headers.get(col): value for col, value in values.items() if headers.get(col)}
                row.clear()


def fast_load_raw_records(path, year, month, stores, by_code, product_by_exact, product_by_norm):
    records = []
    aggregates = defaultdict(lambda: {"qty": 0, "ids": [], "rows": []})
    label = core.month_label(year, month)
    for excel_row, row in enumerate(iter_xlsx_rows(path), start=2):
        raw_code = core.as_int(row.get("store_code"))
        raw_name = str(row.get("item_product_name") or "").strip()
        if raw_code is None or not raw_name:
            continue
        qty = n(row.get("value"))
        report_name = core.norm_name(raw_name)
        master = product_by_exact.get(raw_name) or product_by_norm.get(report_name) or {}
        store_key = core.resolve_raw_store(raw_code, row.get("outlet_name"), stores, by_code)
        key = (store_key, report_name) if store_key else (("UNMAPPED", raw_code, str(row.get("outlet_name") or "")), report_name)
        record = {
            "Status": "Review",
            "month": label,
            "Report Month": label,
            "source_file": str(path),
            "excel_row": excel_row,
            "id": core.as_int(row.get("id")),
            "entry_id": core.as_int(row.get("entry_id")),
            "record_id": core.as_int(row.get("record_id")),
            "entry_date": row.get("entry_date"),
            "user_name": row.get("user_name"),
            "outlet_id": core.as_int(row.get("outlet_id")),
            "outlet_name": row.get("outlet_name"),
            "store_code": raw_code,
            "store_key": store_key,
            "product_name_raw": raw_name,
            "product_name": report_name,
            "p_group": master.get("P_group"),
            "category": master.get("Category"),
            "p_code": master.get("P_code"),
            "product_sku": master.get("Attribute Group") or master.get("label") or row.get("item_label"),
            "qty": qty,
        }
        records.append(record)
        aggregates[key]["qty"] += qty
        aggregates[key]["ids"].append(record["id"])
        aggregates[key]["rows"].append(record)
    return records, aggregates


def fast_load_all_raw_data(raw_files, stores, by_code, product_by_exact, product_by_norm):
    records_by_month = {}
    aggregates_by_month = {}
    for (year, month), path in raw_files.items():
        print(f"loading {core.month_label(year, month)}: {path.name}", flush=True)
        records, aggregates = fast_load_raw_records(path, year, month, stores, by_code, product_by_exact, product_by_norm)
        records_by_month[(year, month)] = records
        aggregates_by_month[(year, month)] = aggregates
    return records_by_month, aggregates_by_month


def fast_sell_in_by_month(months, canonical, product_by_exact, product_by_norm):
    sell_in_all, _exceptions = core.load_sell_in(months, canonical)
    by_month = defaultdict(lambda: defaultdict(lambda: [0, 0, 0]))
    for (year, month, store_key, product), values in sell_in_all.items():
        report_name = core.norm_name(product)
        if product_by_exact.get(product) or product_by_norm.get(report_name):
            target = by_month[(year, month)][(store_key, report_name)]
            for index, value in enumerate(values):
                target[index] += value
    return by_month


def v3_summary_by_month(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Summary By Month"]
    headers = core.header_row(ws, 1)
    idx = {header: index for index, header in enumerate(headers)}
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[idx["Month"]]
        if label:
            data[label] = {
                "stock_baht": n(row[idx["Total Stock (Baht)"]]),
                "sell_baht": n(row[idx["Total Sell-in (Baht)"]]),
                "cvd": n(row[idx["CVD"]]),
            }
    wb.close()
    return data


def existing_v3_raw_file():
    for path in hist.V3_RAW_FALLBACKS:
        if path.exists():
            return path
    raise FileNotFoundError("No V.3 raw data workbook found for comparison.")


def write_sheet(wb, name, headers, rows, fill="1F4E78"):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    core.style_sheet(ws, fill)
    core.set_widths(ws, {1: 16, 2: 24, 3: 24, 10: 70, 11: 70})
    return ws


def main():
    raw_files = hist.raw_files_by_month()
    months = sorted(raw_files, key=hist.month_sort_key)
    stores, by_code, canonical = core.load_store_master()
    product_by_exact, product_by_norm, _pcode, _master = core.load_product_master()
    rsp_by_name = core.load_rsp()

    records_by_month, aggregates_by_month = fast_load_all_raw_data(raw_files, stores, by_code, product_by_exact, product_by_norm)
    print("loading sell-in once for all months", flush=True)
    sell_in_by_month = fast_sell_in_by_month(months, canonical, product_by_exact, product_by_norm)
    _corrections_by_month, issues = hist.build_historical_corrections(months, records_by_month, aggregates_by_month, sell_in_by_month, 6, stores, "TT")

    by_month = defaultdict(lambda: {"count": 0, "old_qty": 0, "new_qty": 0, "stock_baht_diff": 0})
    by_pgroup = defaultdict(lambda: {"count": 0, "old_qty": 0, "new_qty": 0, "stock_baht_diff": 0})
    detail_rows = []
    for issue in issues:
        store_key = None
        # Issue rows do not keep store_key, so use report output fields and current qty only for impact.
        label = issue["Report Month"]
        product_name = issue["Product Name"]
        old_qty = n(issue["Current Qty"])
        new_qty = n(issue["Suggested Correct Qty"])
        rsp = product_rsp(label, product_name, product_by_norm, rsp_by_name)
        stock_baht_diff = (new_qty - old_qty) * rsp
        month_data = by_month[label]
        month_data["count"] += 1
        month_data["old_qty"] += old_qty
        month_data["new_qty"] += new_qty
        month_data["stock_baht_diff"] += stock_baht_diff
        pg_key = (label, issue["P_Group"], issue["Category"])
        pg_data = by_pgroup[pg_key]
        pg_data["count"] += 1
        pg_data["old_qty"] += old_qty
        pg_data["new_qty"] += new_qty
        pg_data["stock_baht_diff"] += stock_baht_diff
        detail_rows.append([
            label,
            issue["id"],
            issue["entry_id"],
            issue["record_id"],
            issue["store_code"],
            issue["outlet_name"],
            issue["P_Group"],
            issue["Category"],
            product_name,
            issue["Issue Type"],
            issue["Severity"],
            old_qty,
            new_qty,
            new_qty - old_qty,
            rsp,
            stock_baht_diff,
            issue["Evidence"],
        ])

    v3_path = existing_v3_raw_file()
    v3 = v3_summary_by_month(v3_path)
    month_rows = []
    for label in sorted(v3):
        diff = by_month[label]["stock_baht_diff"]
        v3_stock = v3[label]["stock_baht"]
        v4_stock = v3_stock + diff
        v3_sell = v3[label]["sell_baht"]
        v3_cvd = v3[label]["cvd"]
        v4_cvd = core.cvd(v4_stock, v3_sell)
        month_rows.append([
            label,
            by_month[label]["count"],
            by_month[label]["old_qty"],
            by_month[label]["new_qty"],
            by_month[label]["new_qty"] - by_month[label]["old_qty"],
            v3_stock,
            v4_stock,
            diff,
            diff / v3_stock if v3_stock else None,
            v3_sell,
            v3_cvd,
            v4_cvd,
            v4_cvd - v3_cvd,
        ])

    pgroup_rows = [
        [label, pg, cat, data["count"], data["old_qty"], data["new_qty"], data["new_qty"] - data["old_qty"], data["stock_baht_diff"]]
        for (label, pg, cat), data in sorted(by_pgroup.items())
    ]

    wb = Workbook()
    ws = wb.active
    wb.remove(ws)
    write_sheet(wb, "V3 vs V4 By Month", ["Month", "Correction Rows", "Old Qty Sum", "New Qty Sum", "Qty Diff", "V3 Stock Baht", "V4 Simulated Stock Baht", "Stock Baht Diff", "Stock Baht Diff %", "Sell-in Baht", "V3 CVD", "V4 Simulated CVD", "CVD Diff"], month_rows)
    write_sheet(wb, "Impact By P_Group", ["Month", "P_Group", "Category", "Correction Rows", "Old Qty Sum", "New Qty Sum", "Qty Diff", "Stock Baht Diff"], pgroup_rows)
    write_sheet(wb, "Correction Detail", ["Month", "id", "entry_id", "record_id", "store_code", "outlet_name", "P_Group", "Category", "Product Name", "Issue Type", "Severity", "Old Qty", "New Qty", "Qty Diff", "RSP", "Stock Baht Diff", "Evidence"], detail_rows, "9C6500")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(json.dumps({
        "output": str(OUTPUT),
        "v3_comparison_file": str(v3_path),
        "months": [core.month_label(*month) for month in months],
        "issues": len(issues),
        "stock_baht_diff_total": sum(row[7] for row in month_rows),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
