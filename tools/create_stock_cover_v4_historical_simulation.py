from __future__ import annotations

import argparse
import json
import shutil
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook


SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import build_stock_cover_v3_all_months as core  # noqa: E402
import run_stock_cover_monthly as monthly  # noqa: E402
import validate_stock_cover_monthly_inputs as validator  # noqa: E402


OUTPUT_DIR = core.REPORT_DIR / "outputs" / "stock_cover_monthly"
VALIDATION_DIR = core.REPORT_DIR / "outputs" / "validation"
EDITED_SOURCE_DIR = VALIDATION_DIR / "historical_edited_raw_source"
V3_RAW_FILE = core.REPORT_DIR / "Raw Data for Stock Cover Day Report May 2026.xlsx"
V3_RAW_FALLBACKS = [
    V3_RAW_FILE,
    core.REPORT_DIR / "outputs" / "stock_cover_monthly" / "Raw Data Stock Cover Day May 2026 - test.xlsx",
]
V4_OUTPUT = OUTPUT_DIR / "Raw data Stock Cover Day - V.4 Historical Edit Simulation.xlsx"
V4_IMPACT_OUTPUT = VALIDATION_DIR / "Historical Edit V4 Impact Summary.xlsx"


def month_sort_key(item):
    year, month = item
    return year * 12 + month


def n(value):
    if value is None:
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def clean_qty(value):
    value = n(value)
    if value < 0:
        value = 0
    return int(value) if value.is_integer() else value


def raw_files_by_month():
    files = {}
    for raw_dir in monthly.RAW_STOCK_DIRS:
        if not raw_dir.exists():
            continue
        for path in raw_dir.glob("*.xlsx"):
            parsed = core.raw_month_from_name(path.name)
            if not parsed:
                continue
            priority = 1 if core.re.search(r"\b(28|29|30|31)\b", path.name) else 0
            current = files.get(parsed)
            if current is None or (priority, path.name) > (current[0], current[1].name):
                files[parsed] = (priority, path)
    return {key: value[1] for key, value in sorted(files.items(), key=lambda item: month_sort_key(item[0]))}


def previous_months(months, current, history_months):
    current_index = months.index(current)
    return months[max(0, current_index - history_months):current_index]


def load_all_raw_data(raw_files, stores, by_code, product_by_exact, product_by_norm):
    records_by_month = {}
    aggregates_by_month = {}
    for (year, month), path in raw_files.items():
        print(f"loading {core.month_label(year, month)}: {path.name}")
        records, aggregates = validator.load_raw_records(path, year, month, stores, by_code, product_by_exact, product_by_norm)
        records_by_month[(year, month)] = records
        aggregates_by_month[(year, month)] = aggregates
    return records_by_month, aggregates_by_month


def get_sell_in_by_month(months, canonical, product_by_exact, product_by_norm):
    sell_in_by_month = {}
    for year, month in months:
        sell_in_by_month[(year, month)] = validator.load_sell_in_baseline(year, month, canonical, product_by_exact, product_by_norm)
    return sell_in_by_month


def build_historical_corrections(months, records_by_month, aggregates_by_month, sell_in_by_month, history_months, stores=None, channel_filter=None):
    corrections_by_month = defaultdict(dict)
    issues = []
    for current in months:
        prior_months = previous_months(months, current, history_months)
        label = core.month_label(*current)
        print(f"validating {label}: prior_months={len(prior_months)}")
        if not prior_months:
            continue

        for record in records_by_month[current]:
            if channel_filter:
                store_key = record.get("store_key")
                if not store_key or not stores or stores.get(store_key, {}).get("Channel") != channel_filter:
                    continue
            key = (record["store_key"], record["product_name"]) if record["store_key"] else (("UNMAPPED", record["store_code"], str(record["outlet_name"] or "")), record["product_name"])
            previous_values = [
                aggregates_by_month[month].get(key, {}).get("qty")
                for month in prior_months
                if aggregates_by_month[month].get(key, {}).get("qty") is not None
            ]
            sell_values = sell_in_by_month[current].get((record["store_key"], record["product_name"]), [0, 0, 0])
            current_qty = record["qty"]
            if current_qty < 0:
                issue = validator.issue_row(record, "Negative stock qty", "High", 0, "Stock qty is negative.", previous_values, sell_values)
                corrections_by_month[current][int(record["id"])] = issue
                issues.append(issue)
                continue

            hist_median = validator.median(previous_values)
            hist_avg = validator.avg(previous_values)
            hist_max = max(previous_values) if previous_values else None
            sell_avg = sum(sell_values) / 3 if sell_values else 0

            issue = None
            if hist_max is not None and hist_max > 0 and current_qty >= max(24, hist_max * 4):
                suggested = hist_median if hist_median is not None else previous_values[-1]
                issue = validator.issue_row(
                    record,
                    "Stock spike vs history",
                    "High",
                    suggested,
                    f"Current qty {current_qty:g} is at least 4x historical max {hist_max:g}; possible pack-size/key-in issue.",
                    previous_values,
                    sell_values,
                )
            elif hist_median is not None and hist_median > 0 and current_qty >= max(18, hist_median * 5):
                issue = validator.issue_row(
                    record,
                    "Stock spike vs historical median",
                    "Medium",
                    hist_median,
                    f"Current qty {current_qty:g} is at least 5x historical median {hist_median:g}.",
                    previous_values,
                    sell_values,
                )
            elif sell_avg > 0 and current_qty >= max(24, sell_avg * 2.5) and (hist_avg is None or current_qty >= max(hist_avg * 3, 24)):
                issue = validator.issue_row(
                    record,
                    "Stock high vs sell-in",
                    "Medium",
                    hist_median if hist_median is not None else (previous_values[-1] if previous_values else None),
                    f"Current qty {current_qty:g} is high vs 3-month average sell-in {sell_avg:g}.",
                    previous_values,
                    sell_values,
                )
            if issue and issue["Suggested Correct Qty"] is not None:
                corrections_by_month[current][int(record["id"])] = issue
                issues.append(issue)
    return corrections_by_month, issues


def apply_corrections(raw_files, corrections_by_month):
    EDITED_SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    edited_files = {}
    applied = []
    for (year, month), raw_path in raw_files.items():
        label = core.month_label(year, month)
        out_path = EDITED_SOURCE_DIR / f"{raw_path.stem} - edit.xlsx"
        shutil.copy2(raw_path, out_path)
        corrections = corrections_by_month.get((year, month), {})
        if not corrections:
            edited_files[(year, month)] = out_path
            continue

        wb = openpyxl.load_workbook(out_path, read_only=False, data_only=False)
        ws = wb["Campaign 299"]
        headers = core.header_row(ws, 1)
        idx = {header: index for index, header in enumerate(headers)}
        id_col = idx["id"] + 1
        value_col = idx["value"] + 1
        content_col = idx["content"] + 1
        for row_num in range(2, ws.max_row + 1):
            raw_id = ws.cell(row_num, id_col).value
            if raw_id is None:
                continue
            try:
                raw_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            correction = corrections.get(raw_id)
            if not correction:
                continue
            old_value = ws.cell(row_num, value_col).value
            new_value = clean_qty(correction["Suggested Correct Qty"])
            ws.cell(row_num, value_col).value = new_value
            ws.cell(row_num, content_col).value = str(new_value)
            applied.append({
                "Month": label,
                "id": raw_id,
                "entry_id": correction.get("entry_id"),
                "record_id": correction.get("record_id"),
                "store_code": correction.get("store_code"),
                "outlet_name": correction.get("outlet_name"),
                "Product Name": correction.get("Product Name"),
                "Issue Type": correction.get("Issue Type"),
                "Severity": correction.get("Severity"),
                "Old Qty": old_value,
                "New Qty": new_value,
                "Evidence": correction.get("Evidence"),
                "Edited Raw Source File": str(out_path),
            })
        wb.save(out_path)
        wb.close()
        edited_files[(year, month)] = out_path
    return edited_files, applied


def add_agg(target, row, idx):
    stock = row[idx["#Total Stock"]] or 0
    stock_baht = row[idx["Total Stock (Baht)"]] or 0
    sell = row[idx["Sell in (1+2+3)/3"]] or 0
    sell_baht = row[idx[" Total Sell in (Baht) "]] or 0
    target["rows"] += 1
    target["stores"].add((row[idx[monthly.CODE_HEADER]], row[idx["Store Name"]]))
    target["skus"].add(row[idx["Product Name"]])
    target["stock"] += stock
    target["stock_baht"] += stock_baht
    target["sell"] += sell
    target["sell_baht"] += sell_baht


def build_v4_workbook(raw_files, output, applied):
    months = sorted(raw_files, key=month_sort_key)
    stores, by_code, canonical = core.load_store_master()
    product_by_exact, product_by_norm, pcode_by_name, product_master_by_name = core.load_product_master()
    rsp_by_name = core.load_rsp()
    sell_in, sell_in_exceptions = core.load_sell_in(months, canonical)
    selected_store_keys = [key for key, store in stores.items() if store.get("Channel") == "TT"]
    raw_idx = {header: index for index, header in enumerate(core.HEADERS)}

    wb = Workbook(write_only=False)
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    ws_raw.append(core.HEADERS)

    month_summary = defaultdict(lambda: {"rows": 0, "stores": set(), "skus": set(), "stock": 0, "stock_baht": 0, "sell": 0, "sell_baht": 0, "raw_file": None, "stock_exception_qty": 0})
    pgroup_summary = defaultdict(lambda: {"rows": 0, "stores": set(), "skus": set(), "stock": 0, "stock_baht": 0, "sell": 0, "sell_baht": 0})

    for year, month in months:
        path = raw_files[(year, month)]
        label = core.month_label(year, month)
        print(f"building V4 raw data {label}: {path.name}")
        products, stock, _raw_store_keys, stock_exceptions = core.load_raw_month(year, month, path, stores, by_code, product_by_exact, product_by_norm, rsp_by_name)
        month_sell_in = core.resolve_month_sell_in(year, month, products, sell_in, rsp_by_name, pcode_by_name, product_master_by_name, label)
        month_summary[label]["raw_file"] = path.name
        month_summary[label]["stock_exception_qty"] = sum(value["qty"] for value in stock_exceptions.values())

        for store_key in sorted(selected_store_keys, key=lambda item: (item[0], item[1])):
            store_row = stores[store_key]
            store_code = store_row["_store_code"]
            store_name = store_row.get("Store Name (SCVD)") or store_row.get("Store Name (CloudViu)") or store_row.get("Store Name (Group)")
            store_group = store_row.get("Store Name (Group)") or store_name
            for product_name in sorted(products):
                product = products[product_name]
                current = stock.get((store_key, product_name), {}).get("current", 0)
                new = stock.get((store_key, product_name), {}).get("new", 0)
                total = current + new
                rsp = product["rsp"] or 0
                sell_data = month_sell_in.get((store_key, product_name))
                sell_values = sell_data["qty"] if sell_data else [0, 0, 0]
                sell_avg = sum(sell_values) / 3
                sell_baht = sell_data["baht"] if sell_data else 0
                row = [
                    store_row.get("Area"),
                    store_row.get("Sales Region (TT)"),
                    product["P_Group"],
                    product["Category"],
                    product["P_Code"],
                    product["Product SKU"],
                    product["Product Name"],
                    product["Size"],
                    product["Brand"],
                    product["Segment"],
                    product["Group"],
                    store_code,
                    store_row.get("Group Code"),
                    store_row.get("Channel"),
                    store_row.get("Type"),
                    store_name,
                    store_row.get("Province (EN)"),
                    core.report_date(year, month),
                    store_code,
                    product["product_id"],
                    f"{store_code}{product['Product Name']}",
                    f"{product['Product SKU']}{store_group}",
                    rsp,
                    current,
                    new,
                    total,
                    total * rsp,
                    sell_values[0],
                    sell_values[1],
                    sell_values[2],
                    sell_avg,
                    sell_baht,
                    label,
                    year,
                ]
                ws_raw.append(row)
                add_agg(month_summary[label], row, raw_idx)
                add_agg(pgroup_summary[(label, row[raw_idx["P_Group"]], row[raw_idx["Category"]])], row, raw_idx)

    core.style_sheet(ws_raw)
    core.set_widths(ws_raw, {3: 24, 4: 24, 6: 28, 7: 42, 16: 34, 21: 42, 22: 42})

    month_rows = [
        [label, data["raw_file"], data["rows"], len(data["stores"]), len(data["skus"]), data["stock"], data["stock_baht"], data["sell"], data["sell_baht"], core.cvd(data["stock_baht"], data["sell_baht"]), data["stock_exception_qty"]]
        for label, data in sorted(month_summary.items())
    ]
    core.write_sheet(wb, "Summary By Month", ["Month", "Raw File", "Rows", "Stores", "SKUs", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD", "Stock Exception Qty"], month_rows, {2: 36})

    pgroup_rows = [
        [label, pg, cat, data["rows"], len(data["stores"]), len(data["skus"]), data["stock"], data["stock_baht"], data["sell"], data["sell_baht"], core.cvd(data["stock_baht"], data["sell_baht"])]
        for (label, pg, cat), data in sorted(pgroup_summary.items())
    ]
    core.write_sheet(wb, "Summary By P_Group", ["Month", "P_Group", "Category", "Rows", "Stores", "SKUs", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"], pgroup_rows, {2: 24, 3: 24})

    correction_headers = ["Month", "id", "entry_id", "record_id", "store_code", "outlet_name", "Product Name", "Issue Type", "Severity", "Old Qty", "New Qty", "Evidence", "Edited Raw Source File"]
    correction_rows = [[item.get(header) for header in correction_headers] for item in applied]
    core.write_sheet(wb, "Corrections Applied", correction_headers, correction_rows, {6: 34, 7: 42, 12: 70, 13: 70}, "9C6500")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return month_rows, pgroup_rows


def read_summary_by_month(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Summary By Month"]
    headers = core.header_row(ws, 1)
    idx = {header: index for index, header in enumerate(headers)}
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[idx["Month"]]
        if not label:
            continue
        data[label] = {
            "stock_baht": n(row[idx["Total Stock (Baht)"]]),
            "sell_baht": n(row[idx["Total Sell-in (Baht)"]]),
            "cvd": n(row[idx["CVD"]]),
        }
    wb.close()
    return data


def write_impact_summary(v3_path, v4_path, applied, output):
    v3 = read_summary_by_month(v3_path)
    v4 = read_summary_by_month(v4_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "V3 vs V4 By Month"
    headers = ["Month", "V3 Stock Baht", "V4 Stock Baht", "Stock Baht Diff", "Stock Baht Diff %", "V3 Sell-in Baht", "V4 Sell-in Baht", "Sell-in Diff", "V3 CVD", "V4 CVD", "CVD Diff"]
    ws.append(headers)
    for label in sorted(set(v3) | set(v4)):
        v3_stock = v3.get(label, {}).get("stock_baht", 0)
        v4_stock = v4.get(label, {}).get("stock_baht", 0)
        v3_sell = v3.get(label, {}).get("sell_baht", 0)
        v4_sell = v4.get(label, {}).get("sell_baht", 0)
        v3_cvd = v3.get(label, {}).get("cvd", 0)
        v4_cvd = v4.get(label, {}).get("cvd", 0)
        ws.append([label, v3_stock, v4_stock, v4_stock - v3_stock, (v4_stock - v3_stock) / v3_stock if v3_stock else None, v3_sell, v4_sell, v4_sell - v3_sell, v3_cvd, v4_cvd, v4_cvd - v3_cvd])
    core.style_sheet(ws)
    core.set_widths(ws, {1: 14, 2: 18, 3: 18, 4: 18, 5: 16})

    ws2 = wb.create_sheet("Correction Summary")
    by_month = defaultdict(lambda: {"count": 0, "old": 0, "new": 0})
    for item in applied:
        month = item["Month"]
        by_month[month]["count"] += 1
        by_month[month]["old"] += n(item["Old Qty"])
        by_month[month]["new"] += n(item["New Qty"])
    ws2.append(["Month", "Correction Rows", "Old Qty Sum", "New Qty Sum", "Qty Diff"])
    for label, data in sorted(by_month.items()):
        ws2.append([label, data["count"], data["old"], data["new"], data["new"] - data["old"]])
    core.style_sheet(ws2, "9C6500")
    core.set_widths(ws2, {1: 14, 2: 18, 3: 18, 4: 18, 5: 18})

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main():
    parser = argparse.ArgumentParser(description="Build V.4 historical stock correction simulation and compare to V.3 raw data.")
    parser.add_argument("--history-months", type=int, default=6, help="Prior months used for each month's validation baseline.")
    args = parser.parse_args()

    raw_files = raw_files_by_month()
    months = sorted(raw_files, key=month_sort_key)
    stores, by_code, canonical = core.load_store_master()
    product_by_exact, product_by_norm, _pcode, _master = core.load_product_master()
    records_by_month, aggregates_by_month = load_all_raw_data(raw_files, stores, by_code, product_by_exact, product_by_norm)
    sell_in_by_month = get_sell_in_by_month(months, canonical, product_by_exact, product_by_norm)
    corrections_by_month, issues = build_historical_corrections(months, records_by_month, aggregates_by_month, sell_in_by_month, args.history_months, stores, "TT")
    edited_files, applied = apply_corrections(raw_files, corrections_by_month)
    build_v4_workbook(edited_files, V4_OUTPUT, applied)
    write_impact_summary(V3_RAW_FILE, V4_OUTPUT, applied, V4_IMPACT_OUTPUT)
    print(json.dumps({
        "v4_output": str(V4_OUTPUT),
        "impact_summary": str(V4_IMPACT_OUTPUT),
        "months": [core.month_label(*month) for month in months],
        "issues_flagged": len(issues),
        "corrections_applied": len(applied),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
