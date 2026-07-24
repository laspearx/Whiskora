from __future__ import annotations

import argparse
import calendar
import json
import statistics
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import build_stock_cover_v3_all_months as core  # noqa: E402
import run_stock_cover_monthly as monthly  # noqa: E402


OUTPUT_ROOT = core.REPORT_DIR / "outputs"
RAW_SHEET = "Campaign 299"


def parse_report_month(value: str) -> tuple[int, int]:
    return monthly.parse_report_month(value)


def month_key(year: int, month: int) -> int:
    return year * 12 + month


def add_months(year: int, month: int, delta: int) -> tuple[int, int]:
    return core.add_months(year, month, delta)


def n(value):
    if value is None:
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def median(values):
    cleaned = [value for value in values if value is not None]
    if not cleaned:
        return None
    return statistics.median(cleaned)


def avg(values):
    cleaned = [value for value in values if value is not None]
    if not cleaned:
        return None
    return sum(cleaned) / len(cleaned)


def output_default(report_year: int, report_month: int) -> Path:
    folder = OUTPUT_ROOT / "validation"
    return folder / f"Stock Raw Data Validation - {calendar.month_abbr[report_month]} {report_year}.xlsx"


def raw_files_by_month():
    files = {}
    for raw_dir in monthly.RAW_STOCK_DIRS:
        if not raw_dir.exists():
            continue
        for path in raw_dir.glob("*.xlsx"):
            parsed = core.raw_month_from_name(path.name)
            if not parsed:
                continue
            priority = 1 if core.re.search(r"\b(28|29|30|31)\b", path.name) else 0
            current = files.get(parsed)
            if current is None or (priority, path.name) > (current[0], current[1].name):
                files[parsed] = (priority, path)
    return {key: value[1] for key, value in files.items()}


def source_months(report_year: int, report_month: int, history_months: int):
    wanted = []
    for delta in range(-history_months, 1):
        wanted.append(add_months(report_year, report_month, delta))
    return wanted


def load_product_indexes():
    product_by_exact, product_by_norm, _pcode_by_name, _master_by_name = core.load_product_master()
    return product_by_exact, product_by_norm


def product_meta(raw_name, product_by_exact, product_by_norm):
    report_name = core.norm_name(raw_name)
    master = product_by_exact.get(raw_name) or product_by_norm.get(report_name) or {}
    return report_name, master


def load_raw_records(path: Path, year: int, month: int, stores, by_code, product_by_exact, product_by_norm):
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb[RAW_SHEET]
    headers = core.header_row(ws, 1)
    idx = {header: index for index, header in enumerate(headers)}

    records = []
    aggregates = defaultdict(lambda: {"qty": 0, "ids": [], "rows": []})
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_code = core.as_int(row[idx["store_code"]])
        raw_name = str(row[idx["item_product_name"]] or "").strip()
        if raw_code is None or not raw_name:
            continue
        qty = n(row[idx["value"]])
        report_name, master = product_meta(raw_name, product_by_exact, product_by_norm)
        store_key = core.resolve_raw_store(raw_code, row[idx["outlet_name"]], stores, by_code)
        key = (store_key, report_name) if store_key else (("UNMAPPED", raw_code, str(row[idx["outlet_name"]] or "")), report_name)
        record = {
            "month": core.month_label(year, month),
            "year": year,
            "month_num": month,
            "source_file": str(path),
            "excel_row": excel_row,
            "id": row[idx.get("id")] if "id" in idx else None,
            "entry_id": row[idx.get("entry_id")] if "entry_id" in idx else None,
            "record_id": row[idx.get("record_id")] if "record_id" in idx else None,
            "entry_date": row[idx.get("entry_date")] if "entry_date" in idx else None,
            "user_name": row[idx.get("user_name")] if "user_name" in idx else None,
            "outlet_id": row[idx.get("outlet_id")] if "outlet_id" in idx else None,
            "outlet_name": row[idx.get("outlet_name")] if "outlet_name" in idx else None,
            "store_code": raw_code,
            "store_key": store_key,
            "product_name_raw": raw_name,
            "product_name": report_name,
            "p_group": master.get("P_group"),
            "category": master.get("Category"),
            "p_code": master.get("P_code"),
            "product_sku": master.get("Attribute Group") or master.get("label") or row[idx.get("item_label")],
            "qty": qty,
            "content": row[idx.get("content")] if "content" in idx else None,
        }
        records.append(record)
        aggregates[key]["qty"] += qty
        aggregates[key]["ids"].append(record["id"])
        aggregates[key]["rows"].append(record)

    wb.close()
    return records, aggregates


def load_sell_in_baseline(report_year: int, report_month: int, canonical, product_by_exact, product_by_norm):
    months = [(report_year, report_month)]
    sell_in, _exceptions = core.load_sell_in(months, canonical)
    by_key = defaultdict(lambda: [0, 0, 0])
    for (_year, _month, store_key, product), values in sell_in.items():
        report_name, _master = product_meta(product, product_by_exact, product_by_norm)
        target = by_key[(store_key, report_name)]
        for index, value in enumerate(values):
            target[index] += value
    return by_key


def issue_row(record, issue_type, severity, suggested_qty, evidence, previous_values, sell_values):
    current = record["qty"]
    return {
        "Status": "Review",
        "Severity": severity,
        "Issue Type": issue_type,
        "Report Month": record["month"],
        "Source File": record["source_file"],
        "Excel Row": record["excel_row"],
        "id": record["id"],
        "entry_id": record["entry_id"],
        "record_id": record["record_id"],
        "entry_date": record["entry_date"],
        "user_name": record["user_name"],
        "outlet_id": record["outlet_id"],
        "outlet_name": record["outlet_name"],
        "store_code": record["store_code"],
        "P_Group": record["p_group"],
        "Category": record["category"],
        "P_Code": record["p_code"],
        "Product SKU": record["product_sku"],
        "Product Name": record["product_name"],
        "Raw Product Name": record["product_name_raw"],
        "Current Qty": current,
        "Suggested Correct Qty": suggested_qty,
        "Qty Difference": None if suggested_qty is None else current - suggested_qty,
        "Previous Month Qty": previous_values[-1] if previous_values else None,
        "Historical Median Qty": median(previous_values),
        "Historical Avg Qty": avg(previous_values),
        "Historical Max Qty": max(previous_values) if previous_values else None,
        "Sell-in TT (1)": sell_values[0] if sell_values else 0,
        "Sell-in TT (2)": sell_values[1] if sell_values else 0,
        "Sell-in TT (3)": sell_values[2] if sell_values else 0,
        "Sell-in Avg": sum(sell_values) / 3 if sell_values else 0,
        "Evidence": evidence,
        "Correction Note": "",
    }


def validate(report_year: int, report_month: int, history_months: int, output: Path) -> dict:
    files = raw_files_by_month()
    months = source_months(report_year, report_month, history_months)
    missing = [core.month_label(y, m) for y, m in months if (y, m) not in files]
    if (report_year, report_month) not in files:
        raise FileNotFoundError(f"No current raw stock file found for {report_year}-{report_month:02d}")

    stores, by_code, canonical = core.load_store_master()
    product_by_exact, product_by_norm = load_product_indexes()

    current_records = []
    history_by_key = defaultdict(dict)
    current_aggregates = {}
    for year, month in months:
        records, aggregates = load_raw_records(files[(year, month)], year, month, stores, by_code, product_by_exact, product_by_norm)
        if (year, month) == (report_year, report_month):
            current_records = records
            current_aggregates = aggregates
        for key, data in aggregates.items():
            history_by_key[key][(year, month)] = data["qty"]

    sell_in = load_sell_in_baseline(report_year, report_month, canonical, product_by_exact, product_by_norm)

    issues = []
    for record in current_records:
        key = (record["store_key"], record["product_name"]) if record["store_key"] else (("UNMAPPED", record["store_code"], str(record["outlet_name"] or "")), record["product_name"])
        previous_months = months[:-1]
        previous_values = [history_by_key.get(key, {}).get(month) for month in previous_months if history_by_key.get(key, {}).get(month) is not None]
        sell_values = sell_in.get((record["store_key"], record["product_name"]), [0, 0, 0])
        current_qty = record["qty"]

        if current_qty < 0:
            issues.append(issue_row(record, "Negative stock qty", "High", 0, "Stock qty is negative; raw data export cannot use a negative stock count.", previous_values, sell_values))
            continue

        hist_median = median(previous_values)
        hist_avg = avg(previous_values)
        hist_max = max(previous_values) if previous_values else None
        previous_qty = previous_values[-1] if previous_values else None
        sell_avg = sum(sell_values) / 3 if sell_values else 0

        if hist_max is not None and hist_max > 0 and current_qty >= max(24, hist_max * 4):
            suggested = hist_median if hist_median is not None else previous_qty
            issues.append(
                issue_row(
                    record,
                    "Stock spike vs history",
                    "High",
                    suggested,
                    f"Current qty {current_qty:g} is at least 4x historical max {hist_max:g}; possible pack-size/key-in issue.",
                    previous_values,
                    sell_values,
                )
            )
            continue

        if hist_median is not None and hist_median > 0 and current_qty >= max(18, hist_median * 5):
            issues.append(
                issue_row(
                    record,
                    "Stock spike vs historical median",
                    "Medium",
                    hist_median,
                    f"Current qty {current_qty:g} is at least 5x historical median {hist_median:g}.",
                    previous_values,
                    sell_values,
                )
            )
            continue

        if sell_avg > 0 and current_qty >= max(24, sell_avg * 2.5) and (hist_avg is None or current_qty >= max(hist_avg * 3, 24)):
            issues.append(
                issue_row(
                    record,
                    "Stock high vs sell-in",
                    "Medium",
                    hist_median if hist_median is not None else previous_qty,
                    f"Current qty {current_qty:g} is high vs 3-month average sell-in {sell_avg:g}; check if pack size was entered as pieces/cartons incorrectly.",
                    previous_values,
                    sell_values,
                )
            )

    build_workbook(output, report_year, report_month, issues, current_records, current_aggregates, missing)
    return {
        "output": str(output),
        "report_month": core.month_label(report_year, report_month),
        "raw_records_checked": len(current_records),
        "issue_count": len(issues),
        "high_count": sum(1 for issue in issues if issue["Severity"] == "High"),
        "medium_count": sum(1 for issue in issues if issue["Severity"] == "Medium"),
        "missing_history_months": missing,
    }


def style_sheet(ws, widths=None, header_fill="1F4E78"):
    fill = PatternFill("solid", fgColor=header_fill)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    widths = widths or {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        width = widths.get(header, 14)
        if header in {"Source File", "Product Name", "Raw Product Name", "Evidence", "Correction Note"}:
            width = max(width, 42)
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00" if isinstance(cell.value, float) and not cell.value.is_integer() else "#,##0"


def build_workbook(output: Path, report_year: int, report_month: int, issues, current_records, current_aggregates, missing):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ["Report Month", core.month_label(report_year, report_month), ""],
        ["Raw records checked", len(current_records), "Rows from latest CloudViu stock raw data"],
        ["Store+product groups checked", len(current_aggregates), "Aggregated by mapped store and product"],
        ["Issues flagged", len(issues), "Rows requiring review before final report refresh"],
        ["High severity issues", sum(1 for issue in issues if issue["Severity"] == "High"), "Likely key-in / pack-size / negative qty issues"],
        ["Medium severity issues", sum(1 for issue in issues if issue["Severity"] == "Medium"), "Unusual movement vs stock history or sell-in trend"],
        ["Missing history months", ", ".join(missing), "Missing months reduce baseline accuracy"],
    ]
    ws.append(["Metric", "Value", "Note"])
    for row in summary_rows:
        ws.append(row)
    style_sheet(ws, {"Metric": 28, "Value": 24, "Note": 70})

    headers = [
        "Status", "Severity", "Issue Type", "Report Month", "Source File", "Excel Row",
        "id", "entry_id", "record_id", "entry_date", "user_name", "outlet_id",
        "outlet_name", "store_code", "P_Group", "Category", "P_Code", "Product SKU",
        "Product Name", "Raw Product Name", "Current Qty", "Suggested Correct Qty",
        "Qty Difference", "Previous Month Qty", "Historical Median Qty", "Historical Avg Qty",
        "Historical Max Qty", "Sell-in TT (1)", "Sell-in TT (2)", "Sell-in TT (3)",
        "Sell-in Avg", "Evidence", "Correction Note",
    ]
    ws_issues = wb.create_sheet("Issues")
    ws_issues.append(headers)
    for issue in issues:
        ws_issues.append([issue.get(header) for header in headers])
    style_sheet(ws_issues, {"Status": 12, "Severity": 12, "Issue Type": 26, "id": 14, "outlet_name": 34, "Evidence": 70})

    tracker_headers = [
        "Status", "id", "entry_id", "record_id", "store_code", "outlet_name",
        "Product Name", "Current Qty", "Suggested Correct Qty", "Qty Difference",
        "Evidence", "Correction Note",
    ]
    ws_tracker = wb.create_sheet("Correction Tracker")
    ws_tracker.append(tracker_headers)
    for issue in issues:
        ws_tracker.append([issue.get(header) for header in tracker_headers])
    style_sheet(ws_tracker, {"id": 14, "outlet_name": 34, "Product Name": 42, "Evidence": 70, "Correction Note": 42}, "9C6500")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main():
    parser = argparse.ArgumentParser(description="Validate monthly CloudViu stock raw data before report export.")
    parser.add_argument("--report-month", required=True, type=parse_report_month, help="Report month, e.g. 2026-05, May 2026, or latest")
    parser.add_argument("--history-months", type=int, default=6, help="Number of prior months used as stock baseline.")
    parser.add_argument("--output", type=Path, help="Output validation workbook path.")
    args = parser.parse_args()

    report_year, report_month = args.report_month
    output = args.output or output_default(report_year, report_month)
    try:
        result = validate(report_year, report_month, args.history_months, output)
    except FileNotFoundError as exc:
        print(f"ERROR: {exc}")
        raise SystemExit(1)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
