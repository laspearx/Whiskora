import collections
import openpyxl


BASE = r"C:\Work\Nestle\Report\(5th) Stock Cover Day"
REPORT = BASE + r"\Stock Cover Day Report May 2026.xlsx"
LOGIC = BASE + r"\Stock Cover Day.xlsx"


def header_map(ws):
    return {cell.value: i for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))}


def main():
    wb = openpyxl.load_workbook(REPORT, read_only=True, data_only=True)
    ws = wb["Raw Data"]
    idx = header_map(ws)

    row_count = 0
    products = set()
    products_by_business_key = set()
    keys = set()
    stores = set()
    by_store = collections.Counter()
    by_prod = collections.Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Month"]] != "05_May-26":
            continue
        row_count += 1
        prod = (
            row[idx["product_id"]],
            row[idx["P_Code"]],
            row[idx["Product SKU"]],
            row[idx["Product Name"]],
            row[idx["Category"]],
            row[idx["P_Group"]],
        )
        prod_bkey = (
            row[idx["P_Code"]],
            row[idx["Product SKU"]],
            row[idx["Product Name"]],
            row[idx["Category"]],
            row[idx["P_Group"]],
        )
        store = (row[idx["Nestlé Code"]], row[idx["Store Name"]], row[idx["Type"]])
        products.add(prod)
        products_by_business_key.add(prod_bkey)
        stores.add(store)
        keys.add((store, prod))
        by_store[store] += 1
        by_prod[prod] += 1

    print("REPORT_MAY_ROWS", row_count)
    print("REPORT_MAY_DISTINCT_STORE_PRODUCT_KEYS", len(keys))
    print("REPORT_MAY_STORES", len(stores))
    print("REPORT_MAY_PRODUCTS", len(products))
    print("REPORT_MAY_PRODUCTS_BUSINESS_KEY", len(products_by_business_key))
    print("REPORT_EXPECTED_CROSSJOIN", len(stores) * len(products))
    print("REPORT_MISSING_FROM_FULL_CROSSJOIN", len(stores) * len(products) - len(keys))
    print("STORE_COUNTS", collections.Counter(by_store.values()).most_common())
    print("PRODUCT_COUNTS", collections.Counter(by_prod.values()).most_common())
    print("SAMPLE_PRODUCTS")
    for item in sorted(products)[:15]:
        print(item)
    wb.close()

    wb = openpyxl.load_workbook(LOGIC, read_only=True, data_only=True)
    ws = wb["Store Master"]
    idx = header_map(ws)
    for channel in ["TT", "MT"]:
        stores_sm = []
        group_codes = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx["Channel"]] == channel:
                stores_sm.append((row[idx["Group Code"]], row[idx["Store Name (CloudViu)"]], row[idx["Type"]]))
                group_codes.add(row[idx["Group Code"]])
        print("STORE_MASTER", channel, len(stores_sm))
        print("STORE_MASTER_GROUP_CODES", channel, len(group_codes))
        print(stores_sm[:10])

    ws = wb["Product Master"]
    idx = header_map(ws)
    status_col = "Status"
    status_counts = collections.Counter()
    active = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        status_counts[row[idx[status_col]]] += 1
        if row[idx[status_col]] == "active":
            active.append(row)
    print("PRODUCT_MASTER_STATUS", status_counts)
    print("ACTIVE_ROWS", len(active))
    print("ACTIVE_UNIQUE_PRODUCT_NAME", len(set(r[idx["Product Name"]] for r in active)))
    print("ACTIVE_UNIQUE_NEW_NAME", len(set(r[idx["Product Name (โฉมใหม่)"]] for r in active)))
    print("ACTIVE_UNIQUE_PCODE", len(set(r[idx["P_code"]] for r in active)))
    print("ACTIVE_SAMPLE")
    for row in active[:15]:
        print(
            row[idx["P_code"]],
            row[idx["Product Name"]],
            row[idx["Product Name (โฉมใหม่)"]],
            row[idx["P_group"]],
            row[idx["Category"]],
            row[idx["Product code"]],
        )
    wb.close()


if __name__ == "__main__":
    main()
