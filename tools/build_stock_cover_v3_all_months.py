from collections import defaultdict
from datetime import datetime
from pathlib import Path
import calendar
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\Work\Nestle")
REPORT_DIR = BASE / r"Report\(5th) Stock Cover Day"
RAW_DIR = REPORT_DIR / "Raw Data Stock"
MASTER_FILE = BASE / r"Master\Final Master.xlsx"
RSP_FILE = BASE / r"Master\RSP on month.xlsx"
SELL_IN_QUERY_FILE = BASE / r"Target & Incentive\Sell in Query.xlsx"
REPORT_OUTPUT = REPORT_DIR / "Raw Data for Stock Cover Day Report May 2026.xlsx"

YEAR_START = 2025
YEAR_END = 2026

HEADERS = [
    "Area",
    "New Region",
    "P_Group",
    "Category",
    "P_Code",
    "Product SKU",
    "Product Name",
    "Size",
    "Brand",
    "Segment",
    "Group",
    "Nestlเธฃเธ Code",
    "Nestlเธฃเธ Group Code",
    "Channel",
    "Type",
    "Store Name",
    "Province",
    "Date",
    "outlet_id",
    "product_id",
    "Concat",
    "Concat2",
    " RSP LTP = TT (-1) ",
    "#Stock เน€เธยเน€เธยเน€เธโฌเน€เธเธเน€เธเธ…เน€เธเธ—เน€เธเธ เน€เธยเน€เธยเน€เธเธเน€เธยเน€เธเธ‘เน€เธยเน€เธยเน€เธเธเน€เธยเน€เธเธ‘เน€เธย",
    "#Stock เน€เธยเน€เธยเน€เธโฌเน€เธเธเน€เธเธ…เน€เธเธ—เน€เธเธ เน€เธยเน€เธยเน€เธเธเน€เธยเน€เธเธเน€เธเธเน€เธย",
    "#Total Stock",
    "Total Stock (Baht)",
    "Sell in = TT (1)",
    "Sell in = TT (2)",
    "Sell in = TT (3)",
    "Sell in (1+2+3)/3",
    " Total Sell in (Baht) ",
    "Month",
    "Year",
]


MONTH_ABBR = {name.lower(): index for index, name in enumerate(calendar.month_abbr) if name}


def as_int(value):
    if value is None:
        return None
    text = str(value).strip()
    try:
        return int(float(text))
    except ValueError:
        match = re.search(r"\d+", text)
        return int(match.group(0)) if match else None


def as_number(value):
    if value is None:
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def norm_name(value):
    return str(value or "").replace(" (เน€เธยเน€เธยเน€เธเธเน€เธยเน€เธเธเน€เธเธเน€เธย)", "").replace(" (เนเธเธกเนเธซเธกเน)", "").strip()


def match_key(value):
    return re.sub(r"\s+", "", str(value or "").strip().casefold())


def raw_month_from_name(name):
    match = re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})", name, re.I)
    if not match:
        return None
    year = int(match.group(2))
    month = MONTH_ABBR[match.group(1).lower()]
    if re.search(r"\b1-3\b", name):
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    return year, month


def month_label(year, month):
    return f"{month:02d}_{calendar.month_abbr[month]}-{str(year)[-2:]}"


def sell_in_month_label(year, month):
    return f"{month:02d}.{year}"


def add_months(year, month, delta):
    month += delta
    while month <= 0:
        month += 12
        year -= 1
    while month > 12:
        month -= 12
        year += 1
    return year, month


def report_date(year, month):
    if month == 12:
        return datetime(year + 1, 1, 1)
    return datetime(year, month + 1, 1)


def choose_raw_files():
    selected = {}
    for path in RAW_DIR.glob("*.xlsx"):
        parsed = raw_month_from_name(path.name)
        if not parsed:
            continue
        year, month = parsed
        if year < YEAR_START or year > YEAR_END:
            continue
        priority = 1 if re.search(r"\b(28|29|30|31)\b", path.name) else 0
        current = selected.get((year, month))
        if current is None or priority > current[0] or (priority == current[0] and path.name > current[1].name):
            selected[(year, month)] = (priority, path)
    return [(year, month, data[1]) for (year, month), data in sorted(selected.items())]


def header_row(ws, row_num):
    return [cell.value for cell in next(ws.iter_rows(min_row=row_num, max_row=row_num))]


def dict_rows(ws, header_row_num, first_data_row):
    headers = header_row(ws, header_row_num)
    for values in ws.iter_rows(min_row=first_data_row, values_only=True):
        if all(value is None for value in values):
            continue
        yield {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}


def add_store_instance(stores, by_code, canonical, row, code, preferred_name=None):
    names = [
        preferred_name,
        row.get("Store Name (CloudViu)"),
        row.get("Store Name (SCVD)"),
        row.get("Store Name (Group)"),
        row.get("Store Name (TH)"),
        row.get("Store Name (EN)"),
    ]
    name_key = next((match_key(name) for name in names if match_key(name)), f"store-{code}")
    key = (code, name_key)
    if key not in stores:
        row = dict(row)
        row["_store_key"] = key
        row["_store_code"] = code
        row["_match_names"] = {match_key(name) for name in names if match_key(name)}
        stores[key] = row
        by_code[code].append(key)
        canonical.setdefault(code, key)
    return key


def index_store_alias(by_code, canonical, alias_code, key):
    alias_code = as_int(alias_code)
    if alias_code is None:
        return
    if key not in by_code[alias_code]:
        by_code[alias_code].append(key)
    canonical.setdefault(alias_code, key)


def load_store_master():
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws = wb["Store Master"]
    stores = {}
    by_code = defaultdict(list)
    canonical = {}
    for row in dict_rows(ws, 15, 16):
        code = as_int(row.get("Store Code"))
        if code is not None:
            key = add_store_instance(stores, by_code, canonical, row, code)
            index_store_alias(by_code, canonical, row.get("storeCode"), key)
    wb.close()
    return stores, by_code, canonical


def load_product_master():
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws = wb["Product Master"]
    by_exact = {}
    by_norm = {}
    pcode_by_name = {}
    master_by_name = {}
    for row in dict_rows(ws, 14, 15):
        exact_names = {
            str(row.get("name") or "").strip(),
            str(row.get("Product Name") or "").strip(),
            str(row.get("Product Name (เนเธเธกเนเธซเธกเน)") or "").strip(),
            str(row.get("Product Name (เน€เธยเน€เธยเน€เธเธเน€เธยเน€เธเธเน€เธเธเน€เธย)") or "").strip(),
        }
        pcode = str(row.get("P_code") or "").strip()
        for exact in exact_names:
            if exact:
                by_exact.setdefault(exact, row)
                pcode_by_name.setdefault(norm_name(exact), pcode)
                master_by_name.setdefault(norm_name(exact), row)
        normalized = norm_name(row.get("Product Name (เนเธเธกเนเธซเธกเน)") or row.get("Product Name (เน€เธยเน€เธยเน€เธเธเน€เธยเน€เธเธเน€เธเธเน€เธย)") or row.get("Product Name") or row.get("name"))
        if normalized:
            by_norm.setdefault(normalized, row)
            pcode_by_name.setdefault(normalized, pcode)
            master_by_name.setdefault(normalized, row)
    wb.close()
    return by_exact, by_norm, pcode_by_name, master_by_name


def load_rsp():
    wb = openpyxl.load_workbook(RSP_FILE, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = header_row(ws, 1)
    idx = {h: i for i, h in enumerate(headers)}
    rsp = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        product = norm_name(row[idx["Product Name"]])
        month = row[idx["Month"]]
        if product and month:
            rsp[(str(month), product)] = row[idx["RSP"]] or 0
    wb.close()
    return rsp


def resolve_raw_store(raw_code, outlet_name, stores, by_code):
    raw_name_key = match_key(outlet_name)
    candidates = by_code.get(raw_code, [])
    for key in candidates:
        if raw_name_key and raw_name_key in stores[key].get("_match_names", set()):
            return key
    if len(candidates) == 1:
        return candidates[0]
    return None


def load_raw_month(year, month, path, stores, by_code, product_by_exact, product_by_norm, rsp_by_name):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Campaign 299"]
    if ws.max_row == 1 and ws.max_column == 1:
        ws.reset_dimensions()
    headers = header_row(ws, 1)
    idx = {h: i for i, h in enumerate(headers)}
    label = month_label(year, month)
    products = {}
    stock = defaultdict(lambda: {"current": 0, "new": 0})
    raw_store_keys = set()
    exceptions = defaultdict(lambda: {"rows": 0, "qty": 0})

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_code = as_int(row[idx["store_code"]])
        if raw_code is None:
            continue
        raw_name = str(row[idx["item_product_name"]] or "").strip()
        report_name = norm_name(raw_name)
        if not report_name:
            continue
        master = product_by_exact.get(raw_name) or product_by_norm.get(report_name) or {}
        variant = str(master.get("SPD") or "").strip()
        bucket = "new" if "เน€เธยเน€เธเธเน€เธเธเน€เธย" in raw_name or "เนเธซเธกเน" in raw_name or "เน€เธยเน€เธเธเน€เธเธเน€เธย" in variant or "เนเธซเธกเน" in variant else "current"
        rsp = rsp_by_name.get((label, report_name)) or master.get("marketPrice") or 0
        products.setdefault(
            report_name,
            {
                "P_Group": master.get("P_group"),
                "Category": master.get("Category"),
                "P_Code": master.get("P_code"),
                "Product SKU": master.get("Attribute Group") or master.get("label") or row[idx["item_label"]],
                "Product Name": report_name,
                "Size": master.get("Size"),
                "Brand": master.get("Brand"),
                "Segment": master.get("Group"),
                "Group": master.get("Type"),
                "product_id": master.get("sku"),
                "rsp": rsp,
            },
        )
        qty = as_number(row[idx["value"]])
        store_key = resolve_raw_store(raw_code, row[idx["outlet_name"]], stores, by_code)
        if store_key is None:
            reason = "Store code not in Final Master" if raw_code not in by_code else "Multiple Final Master stores share code and outlet name did not match"
            key = (label, raw_code, row[idx["outlet_id"]], row[idx["outlet_name"]], report_name, reason)
            exceptions[key]["rows"] += 1
            exceptions[key]["qty"] += qty
            continue
        stock[(store_key, report_name)][bucket] += qty
        raw_store_keys.add(store_key)
    wb.close()
    return products, stock, raw_store_keys, exceptions


def load_sell_in(months, canonical):
    wanted_months = defaultdict(list)
    for report_year, report_month in months:
        for offset, delta in enumerate([-2, -1, 0]):
            source_year, source_month = add_months(report_year, report_month, delta)
            wanted_months[sell_in_month_label(source_year, source_month)].append((report_year, report_month, offset))
    wb = openpyxl.load_workbook(SELL_IN_QUERY_FILE, read_only=True, data_only=True)
    ws = wb["Sell in Query"]
    headers = header_row(ws, 1)
    idx = {h: i for i, h in enumerate(headers)}
    sell_in = defaultdict(lambda: [0, 0, 0])
    exceptions = defaultdict(lambda: {"rows": 0, "qty": 0})

    for row in ws.iter_rows(min_row=2, values_only=True):
        source_month = str(row[idx["Cal. year / month"]] or "")
        if source_month not in wanted_months:
            continue
        customer = as_int(row[idx["Customer"]])
        product = norm_name(row[idx["Product Name (CloudViu)"]])
        if customer is None or not product or product.upper() == "N/A":
            continue
        qty = max(as_number(row[idx["PCS."]]), 0)
        store_key = canonical.get(customer)
        if store_key is None:
            for report_year, report_month, _offset in wanted_months[source_month]:
                key = (month_label(report_year, report_month), customer, row[idx["Customer Name"]], product, "Customer code not in Final Master")
                exceptions[key]["rows"] += 1
                exceptions[key]["qty"] += qty
            continue
        for report_year, report_month, offset in wanted_months[source_month]:
            sell_in[(report_year, report_month, store_key, product)][offset] += qty
    wb.close()
    return sell_in, exceptions


def build_products_by_pcode(products):
    by_pcode = defaultdict(list)
    for product_name, product in products.items():
        pcode = str(product.get("P_Code") or "").strip()
        if pcode:
            by_pcode[pcode].append(product_name)
    return by_pcode


def product_rsp(label, product_name, product, rsp_by_name, product_master_by_name):
    if product and product.get("rsp"):
        return product.get("rsp") or 0
    rsp = rsp_by_name.get((label, product_name))
    if rsp:
        return rsp
    master = product_master_by_name.get(product_name) or {}
    return master.get("marketPrice") or 0


def resolve_sell_product(product_name, products, products_by_pcode, pcode_by_name):
    if product_name in products:
        return product_name
    pcode = pcode_by_name.get(product_name)
    if not pcode:
        return None
    candidates = products_by_pcode.get(pcode, [])
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        combined = [candidate for candidate in candidates if " / " in candidate]
        if len(combined) == 1:
            return combined[0]
    return None


def resolve_month_sell_in(year, month, products, sell_in, rsp_by_name, pcode_by_name, product_master_by_name, label):
    products_by_pcode = build_products_by_pcode(products)
    resolved = defaultdict(lambda: {"qty": [0, 0, 0], "baht": 0})
    for (sell_year, sell_month, store_key, sell_product), values in sell_in.items():
        if (sell_year, sell_month) != (year, month):
            continue
        target_product = resolve_sell_product(sell_product, products, products_by_pcode, pcode_by_name)
        if target_product is None:
            continue
        source_product = products.get(sell_product)
        rsp = product_rsp(label, sell_product, source_product, rsp_by_name, product_master_by_name)
        target = resolved[(store_key, target_product)]
        for index, value in enumerate(values):
            target["qty"][index] += value
        target["baht"] += (sum(values) / 3) * rsp
    return resolved


def cvd(stock_baht, sell_baht):
    return stock_baht / sell_baht * 30 if sell_baht else 0


def add_agg(target, row, idx):
    stock = row[idx["#Total Stock"]] or 0
    stock_baht = row[idx["Total Stock (Baht)"]] or 0
    sell = row[idx["Sell in (1+2+3)/3"]] or 0
    sell_baht = row[idx[" Total Sell in (Baht) "]] or 0
    target["rows"] += 1
    target["stores"].add((row[idx["Nestlเธฃเธ Code"]], row[idx["Store Name"]]))
    target["skus"].add(row[idx["Product Name"]])
    target["stock"] += stock
    target["stock_baht"] += stock_baht
    target["sell"] += sell
    target["sell_baht"] += sell_baht


def style_sheet(ws, header_fill="1F4E78"):
    fill = PatternFill("solid", fgColor=header_fill)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False


def set_widths(ws, widths=None, default=14):
    widths = widths or {}
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, default)


def write_sheet(wb, name, headers, data_rows, widths=None, header_fill="1F4E78"):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in data_rows:
        ws.append(row)
    style_sheet(ws, header_fill)
    set_widths(ws, widths)
    return ws


def build_report():
    raw_files = choose_raw_files()
    months = [(year, month) for year, month, _path in raw_files]
    month_order = {month_label(year, month): index for index, (year, month) in enumerate(months)}
    stores, by_code, canonical = load_store_master()
    product_by_exact, product_by_norm, pcode_by_name, product_master_by_name = load_product_master()
    rsp_by_name = load_rsp()
    sell_in, sell_in_exceptions = load_sell_in(months, canonical)
    tt_store_keys = [key for key, store in stores.items() if store.get("Channel") == "TT"]
    raw_idx = {header: i for i, header in enumerate(HEADERS)}

    wb = Workbook(write_only=False)
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    ws_raw.append(HEADERS)

    channel = defaultdict(lambda: {"rows": 0, "stores": set(), "skus": set(), "stock": 0, "stock_baht": 0, "sell": 0, "sell_baht": 0})
    month_summary = defaultdict(lambda: {"rows": 0, "stores": set(), "skus": set(), "stock": 0, "stock_baht": 0, "sell": 0, "sell_baht": 0, "raw_file": None, "raw_stores": 0, "stock_exception_rows": 0, "stock_exception_qty": 0})
    sub_brand = defaultdict(lambda: {"rows": 0, "stores": set(), "skus": set(), "stock": 0, "stock_baht": 0, "sell": 0, "sell_baht": 0})
    store = defaultdict(lambda: {"meta": None, "skus": set(), "stock": 0, "stock_baht": 0, "sell": 0, "sell_baht": 0})
    store_sku = defaultdict(lambda: {"meta": None, "stock": 0, "stock_baht": 0, "sell": 0, "sell_baht": 0})
    stock_exceptions = defaultdict(lambda: {"rows": 0, "qty": 0})

    for year, month, raw_path in raw_files:
        products, stock, raw_store_keys, exceptions = load_raw_month(year, month, raw_path, stores, by_code, product_by_exact, product_by_norm, rsp_by_name)
        label = month_label(year, month)
        ms = month_summary[label]
        ms["raw_file"] = raw_path.name
        ms["raw_stores"] = len(raw_store_keys)
        ms["stock_exception_rows"] += sum(v["rows"] for v in exceptions.values())
        ms["stock_exception_qty"] += sum(v["qty"] for v in exceptions.values())
        for key, value in exceptions.items():
            stock_exceptions[key]["rows"] += value["rows"]
            stock_exceptions[key]["qty"] += value["qty"]
        month_sell_in = resolve_month_sell_in(year, month, products, sell_in, rsp_by_name, pcode_by_name, product_master_by_name, label)

        for store_key in sorted(tt_store_keys, key=lambda item: (item[0], item[1])):
            store_row = stores[store_key]
            store_code = store_row["_store_code"]
            store_name = store_row.get("Store Name (SCVD)") or store_row.get("Store Name (CloudViu)") or store_row.get("Store Name (Group)")
            store_group = store_row.get("Store Name (Group)") or store_name
            for product_name in sorted(products):
                product = products[product_name]
                current = stock.get((store_key, product_name), {}).get("current", 0)
                new = stock.get((store_key, product_name), {}).get("new", 0)
                total = current + new
                rsp = product["rsp"] or 0
                sell_data = month_sell_in.get((store_key, product_name))
                sell_values = sell_data["qty"] if sell_data else [0, 0, 0]
                sell_avg = sum(sell_values) / 3
                sell_baht = sell_data["baht"] if sell_data else 0
                row = [
                    store_row.get("Area"),
                    store_row.get("Sales Region (TT)"),
                    product["P_Group"],
                    product["Category"],
                    product["P_Code"],
                    product["Product SKU"],
                    product["Product Name"],
                    product["Size"],
                    product["Brand"],
                    product["Segment"],
                    product["Group"],
                    store_code,
                    store_row.get("Group Code"),
                    store_row.get("Channel"),
                    store_row.get("Type"),
                    store_name,
                    store_row.get("Province (EN)"),
                    report_date(year, month),
                    store_code,
                    product["product_id"],
                    f"{store_code}{product['Product Name']}",
                    f"{product['Product SKU']}{store_group}",
                    rsp,
                    current,
                    new,
                    total,
                    total * rsp,
                    sell_values[0],
                    sell_values[1],
                    sell_values[2],
                    sell_avg,
                    sell_baht,
                    label,
                    year,
                ]
                ws_raw.append(row)
                add_agg(channel[row[raw_idx["Channel"]]], row, raw_idx)
                add_agg(month_summary[label], row, raw_idx)

                sb_key = (label, row[raw_idx["P_Group"]], row[raw_idx["Category"]])
                add_agg(sub_brand[sb_key], row, raw_idx)

                st_key = (label, store_code, store_name)
                target = store[st_key]
                target["meta"] = [
                    label,
                    row[raw_idx["Area"]],
                    row[raw_idx["New Region"]],
                    store_code,
                    row[raw_idx["Nestlเธฃเธ Group Code"]],
                    row[raw_idx["Channel"]],
                    row[raw_idx["Type"]],
                    store_name,
                    row[raw_idx["Province"]],
                ]
                target["skus"].add(product_name)
                target["stock"] += total
                target["stock_baht"] += total * rsp
                target["sell"] += sell_avg
                target["sell_baht"] += sell_avg * rsp

                ss_key = (label, store_code, store_name, product_name)
                target = store_sku[ss_key]
                target["meta"] = [
                    label,
                    row[raw_idx["Area"]],
                    row[raw_idx["New Region"]],
                    store_code,
                    row[raw_idx["Nestlเธฃเธ Group Code"]],
                    row[raw_idx["Channel"]],
                    row[raw_idx["Type"]],
                    store_name,
                    row[raw_idx["Province"]],
                    row[raw_idx["P_Group"]],
                    row[raw_idx["Category"]],
                    product_name,
                    row[raw_idx["Product SKU"]],
                ]
                target["stock"] += total
                target["stock_baht"] += total * rsp
                target["sell"] += sell_avg
                target["sell_baht"] += sell_avg * rsp
        print(f"processed {label}: products={len(products)} tt_rows_added={len(tt_store_keys) * len(products)}")

    style_sheet(ws_raw)
    set_widths(ws_raw, {3: 24, 4: 24, 6: 28, 7: 42, 16: 34, 21: 42, 22: 42})

    channel_rows = []
    for key, data in sorted(channel.items(), key=lambda item: str(item[0])):
        channel_rows.append([key, data["rows"], len(data["stores"]), len(data["skus"]), data["stock"], data["stock_baht"], data["sell"], data["sell_baht"], cvd(data["stock_baht"], data["sell_baht"])])
    write_sheet(wb, "Summary By Channel", ["Channel", "Rows", "Stores", "SKUs", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"], channel_rows)

    month_rows = []
    for label, data in sorted(month_summary.items(), key=lambda item: month_order[item[0]]):
        month_rows.append([label, data["raw_file"], data["rows"], len(data["stores"]), len(data["skus"]), data["raw_stores"], data["stock"], data["stock_baht"], data["sell"], data["sell_baht"], cvd(data["stock_baht"], data["sell_baht"]), data["stock_exception_rows"], data["stock_exception_qty"]])
    write_sheet(wb, "Summary By Month", ["Month", "Raw File", "Rows", "Stores", "SKUs", "Stores with Raw Stock", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD", "Unmapped Stock Rows", "Unmapped Stock Qty"], month_rows, {2: 28})

    sub_brand_rows = []
    for (label, p_group, category), data in sorted(sub_brand.items(), key=lambda item: (month_order[item[0][0]], str(item[0][1]), str(item[0][2]))):
        sub_brand_rows.append([label, p_group, category, data["rows"], len(data["stores"]), len(data["skus"]), data["stock"], data["stock_baht"], data["sell"], data["sell_baht"], cvd(data["stock_baht"], data["sell_baht"])])
    write_sheet(wb, "Summary By Sub Brand", ["Month", "P_Group", "Category", "Rows", "Stores", "SKUs", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"], sub_brand_rows, {1: 14, 2: 24, 3: 24})

    stock_rows = []
    for (_label, _code, _name), data in sorted(store.items(), key=lambda item: (month_order[item[0][0]], str(item[0][1]), str(item[0][2]))):
        stock_rows.append(data["meta"] + [len(data["skus"]), data["stock"], data["stock_baht"], data["sell"], data["sell_baht"], cvd(data["stock_baht"], data["sell_baht"])])
    write_sheet(wb, "Stock Data", ["Month", "Area", "New Region", "Nestlเธฃเธ Code", "Nestlเธฃเธ Group Code", "Channel", "Type", "Store Name", "Province", "SKUs", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"], stock_rows, {8: 34})

    pivot_store_rows = []
    for (_label, _code, _name), data in sorted(store.items(), key=lambda item: (month_order[item[0][0]], str(item[0][1]), str(item[0][2]))):
        pivot_store_rows.append([data["meta"][0], data["meta"][3], data["meta"][4], data["meta"][7], data["meta"][5], data["meta"][6], data["stock"], data["stock_baht"], data["sell"], data["sell_baht"], cvd(data["stock_baht"], data["sell_baht"])])
    write_sheet(wb, "Pivot by Store", ["Month", "Nestlเธฃเธ Code", "Nestlเธฃเธ Group Code", "Store Name", "Channel", "Type", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"], pivot_store_rows, {4: 34})

    store_sku_rows = []
    for (_label, _code, _name, _product), data in sorted(store_sku.items(), key=lambda item: (month_order[item[0][0]], str(item[0][1]), str(item[0][2]), str(item[0][3]))):
        store_sku_rows.append(data["meta"] + [data["stock"], data["stock_baht"], data["sell"], data["sell_baht"], cvd(data["stock_baht"], data["sell_baht"])])
    write_sheet(wb, "Store by SKU", ["Month", "Area", "New Region", "Nestlเธฃเธ Code", "Nestlเธฃเธ Group Code", "Channel", "Type", "Store Name", "Province", "P_Group", "Category", "Product Name", "Product SKU", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"], store_sku_rows, {8: 34, 12: 42, 13: 28})

    stock_exception_rows = [list(key) + [value["rows"], value["qty"]] for key, value in sorted(stock_exceptions.items(), key=lambda item: (month_order.get(item[0][0], 999),) + tuple(str(part) for part in item[0][1:]))]
    write_sheet(wb, "Stock Mapping Exceptions", ["Month", "Store Code", "Outlet ID", "Outlet Name", "Product Name", "Reason", "Source Rows", "Qty"], stock_exception_rows, {4: 34, 5: 42, 6: 38}, "9C6500")

    sell_exception_rows = [list(key) + [value["rows"], value["qty"]] for key, value in sorted(sell_in_exceptions.items(), key=lambda item: (month_order.get(item[0][0], 999),) + tuple(str(part) for part in item[0][1:]))]
    write_sheet(wb, "Sell-in Mapping Exceptions", ["Month", "Customer Code", "Customer Name", "Product Name", "Reason", "Source Rows", "Qty"], sell_exception_rows, {3: 34, 4: 42, 5: 32}, "9C6500")

    wb.save(REPORT_OUTPUT)
    return {
        "output": str(REPORT_OUTPUT),
        "months": [month_label(y, m) for y, m, _p in raw_files],
        "raw_rows": ws_raw.max_row - 1,
        "stock": sum(row[6] for row in month_rows),
        "stock_baht": sum(row[7] for row in month_rows),
        "sell": sum(row[8] for row in month_rows),
        "sell_baht": sum(row[9] for row in month_rows),
        "stock_exception_qty": sum(row[12] for row in month_rows),
        "sell_in_exception_qty": sum(value["qty"] for value in sell_in_exceptions.values()),
    }


if __name__ == "__main__":
    print(build_report())

