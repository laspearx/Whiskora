from pathlib import Path
import json
import re

import openpyxl


REPORT_DIR = Path(r"C:\Work\Nestle\Report\(5th) Stock Cover Day")
RAW_DIR = REPORT_DIR / "Raw Data Stock"
SELL_IN_FILE = REPORT_DIR / "Stock Cover Day.xlsx"


def raw_month_from_name(name):
    match = re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})", name, re.I)
    if not match:
        return None
    month_names = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
    return f"{int(match.group(2)):04d}-{month_names.index(match.group(1).lower()) + 1:02d}"


def inspect_raw_files():
    result = []
    for path in sorted(RAW_DIR.glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        sheet_title = ws.title
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        mode = "read_only"
        rows = ws.max_row
        cols = ws.max_column
        has_expected = all(h in headers for h in ["store_code", "outlet_name", "item_product_name", "value"])
        wb.close()
        if not has_expected:
            wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
            ws = wb[wb.sheetnames[0]]
            sheet_title = ws.title
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            mode = "normal"
            rows = ws.max_row
            cols = ws.max_column
            has_expected = all(h in headers for h in ["store_code", "outlet_name", "item_product_name", "value"])
            wb.close()
        result.append(
            {
                "file": path.name,
                "month": raw_month_from_name(path.name),
                "sheet": sheet_title,
                "rows": rows,
                "cols": cols,
                "mode": mode,
                "has_expected_headers": has_expected,
            }
        )
    return result


def inspect_sell_in_months():
    wb = openpyxl.load_workbook(SELL_IN_FILE, read_only=True, data_only=True)
    ws = wb["Sell in"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    counts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        month = row[idx["Cal. year / month"]]
        if month is None:
            continue
        month = str(month)
        counts[month] = counts.get(month, 0) + 1
    wb.close()
    return dict(sorted(counts.items()))


def main():
    print(json.dumps({"raw_files": inspect_raw_files(), "sell_in_months": inspect_sell_in_months()}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
