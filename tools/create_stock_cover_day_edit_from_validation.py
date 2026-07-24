from __future__ import annotations

import argparse
import calendar
import json
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import run_stock_cover_monthly as monthly  # noqa: E402
import validate_stock_cover_monthly_inputs as validator  # noqa: E402


REPORT_DIR = monthly.core.REPORT_DIR
VALIDATION_DIR = REPORT_DIR / "outputs" / "validation"
OUTPUT_DIR = REPORT_DIR / "outputs" / "stock_cover_monthly"
EDITED_SOURCE_DIR = REPORT_DIR / "outputs" / "validation" / "edited_raw_source"


def parse_report_month(value: str):
    return monthly.parse_report_month(value)


def month_name(year: int, month: int) -> str:
    return f"{calendar.month_abbr[month]} {year}"


def validation_path(year: int, month: int) -> Path:
    return VALIDATION_DIR / f"Stock Raw Data Validation - {month_name(year, month)}.xlsx"


def edited_source_path(year: int, month: int) -> Path:
    return EDITED_SOURCE_DIR / f"CloudViu Stock Raw Data - edit - {month_name(year, month)}.xlsx"


def output_path(year: int, month: int) -> Path:
    return OUTPUT_DIR / f"Raw data Stock Cover Day - edit - {month_name(year, month)}.xlsx"


def n(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def clean_qty(value):
    value = n(value)
    if value is None:
        return None
    if value < 0:
        value = 0
    return int(value) if float(value).is_integer() else value


def sheet_headers(ws):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return headers, {header: index for index, header in enumerate(headers)}


def load_corrections(path: Path, severities: set[str]):
    if not path.exists():
        raise FileNotFoundError(f"Validation workbook not found: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Correction Tracker"]
    headers, idx = sheet_headers(ws)
    corrections = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = str(row[idx["Status"]] or "").strip().casefold()
        raw_id = row[idx["id"]]
        suggested = clean_qty(row[idx["Suggested Correct Qty"]])
        if raw_id is None or suggested is None:
            continue
        if status not in {"review", "approved", "apply", "yes", ""}:
            continue
        # Severity exists in Issues, not Correction Tracker. It is merged below.
        corrections[int(raw_id)] = {
            "id": int(raw_id),
            "entry_id": row[idx.get("entry_id")],
            "record_id": row[idx.get("record_id")],
            "store_code": row[idx.get("store_code")],
            "outlet_name": row[idx.get("outlet_name")],
            "product_name": row[idx.get("Product Name")],
            "current_qty_tracker": row[idx.get("Current Qty")],
            "suggested_qty": suggested,
            "qty_difference": row[idx.get("Qty Difference")],
            "evidence": row[idx.get("Evidence")],
            "correction_note": row[idx.get("Correction Note")],
        }

    if "Issues" in wb.sheetnames:
        ws_issues = wb["Issues"]
        issue_headers, issue_idx = sheet_headers(ws_issues)
        issue_by_id = {}
        for row in ws_issues.iter_rows(min_row=2, values_only=True):
            raw_id = row[issue_idx["id"]]
            if raw_id is None:
                continue
            issue_by_id[int(raw_id)] = {
                "severity": row[issue_idx["Severity"]],
                "issue_type": row[issue_idx["Issue Type"]],
                "historical_median": row[issue_idx["Historical Median Qty"]],
                "historical_avg": row[issue_idx["Historical Avg Qty"]],
                "sell_in_avg": row[issue_idx["Sell-in Avg"]],
            }
        for raw_id in list(corrections):
            issue = issue_by_id.get(raw_id, {})
            severity = str(issue.get("severity") or "").strip()
            if severities and severity not in severities:
                corrections.pop(raw_id, None)
                continue
            corrections[raw_id].update(issue)

    wb.close()
    return corrections


def apply_corrections_to_raw_source(raw_file: Path, edited_raw_file: Path, corrections: dict[int, dict]):
    edited_raw_file.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(raw_file, edited_raw_file)

    wb = openpyxl.load_workbook(edited_raw_file, read_only=False, data_only=False)
    ws = wb["Campaign 299"]
    headers, idx = sheet_headers(ws)
    id_col = idx["id"] + 1
    value_col = idx["value"] + 1
    content_col = idx["content"] + 1

    applied = []
    for row_num in range(2, ws.max_row + 1):
        raw_id = ws.cell(row_num, id_col).value
        if raw_id is None:
            continue
        try:
            raw_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        correction = corrections.get(raw_id)
        if not correction:
            continue
        old_value = ws.cell(row_num, value_col).value
        new_value = correction["suggested_qty"]
        ws.cell(row_num, value_col).value = new_value
        ws.cell(row_num, content_col).value = str(new_value)
        applied.append({**correction, "excel_row": row_num, "old_value": old_value, "new_value": new_value})

    wb.save(edited_raw_file)
    wb.close()
    return applied


def add_corrections_sheet(workbook_path: Path, applied: list[dict], validation_file: Path, edited_raw_file: Path):
    wb = openpyxl.load_workbook(workbook_path)
    if "Corrections Applied" in wb.sheetnames:
        del wb["Corrections Applied"]
    ws = wb.create_sheet("Corrections Applied")
    headers = [
        "id", "entry_id", "record_id", "excel_row", "store_code", "outlet_name",
        "Product Name", "Issue Type", "Severity", "Old Qty", "New Qty",
        "Historical Median Qty", "Historical Avg Qty", "Sell-in Avg", "Evidence",
        "Validation File", "Edited Raw Source File",
    ]
    ws.append(headers)
    for item in applied:
        ws.append([
            item.get("id"),
            item.get("entry_id"),
            item.get("record_id"),
            item.get("excel_row"),
            item.get("store_code"),
            item.get("outlet_name"),
            item.get("product_name"),
            item.get("issue_type"),
            item.get("severity"),
            item.get("old_value"),
            item.get("new_value"),
            item.get("historical_median"),
            item.get("historical_avg"),
            item.get("sell_in_avg"),
            item.get("evidence"),
            str(validation_file),
            str(edited_raw_file),
        ])
    style_sheet(ws)
    wb.save(workbook_path)
    wb.close()


def style_sheet(ws):
    fill = PatternFill("solid", fgColor="9C6500")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        width = 14
        if header in {"outlet_name", "Product Name"}:
            width = 36
        elif header in {"Evidence", "Validation File", "Edited Raw Source File"}:
            width = 70
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_view.showGridLines = False


def main():
    parser = argparse.ArgumentParser(description="Create edited Stock Cover Day raw data from validation suggested corrections.")
    parser.add_argument("--report-month", required=True, type=parse_report_month, help="Report month, e.g. 2026-05, May 2026, or latest")
    parser.add_argument(
        "--apply-severity",
        default="High,Medium",
        help="Comma-separated severities to apply from validation workbook. Default: High,Medium",
    )
    args = parser.parse_args()

    report_year, report_month = args.report_month
    severities = {item.strip() for item in args.apply_severity.split(",") if item.strip()}
    raw_file = monthly.choose_raw_file(report_year, report_month)
    validation_file = validation_path(report_year, report_month)
    edited_raw_file = edited_source_path(report_year, report_month)
    out_file = output_path(report_year, report_month)

    corrections = load_corrections(validation_file, severities)
    applied = apply_corrections_to_raw_source(raw_file, edited_raw_file, corrections)
    result = monthly.build_monthly_workbook(
        report_year,
        report_month,
        out_file,
        raw_file=edited_raw_file,
        report_type="stock-cover-day",
    )
    add_corrections_sheet(out_file, applied, validation_file, edited_raw_file)
    result.update(
        {
            "validation_file": str(validation_file),
            "edited_raw_source_file": str(edited_raw_file),
            "corrections_loaded": len(corrections),
            "corrections_applied": len(applied),
            "apply_severity": sorted(severities),
        }
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
