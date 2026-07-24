from pathlib import Path
import zipfile
import openpyxl


FILE = Path(r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Raw Data Stock\Stock 31 May 2026.xlsx")


def inspect_openpyxl():
    wb = openpyxl.load_workbook(FILE, read_only=False, data_only=True)
    print("openpyxl sheets:", wb.sheetnames)
    for ws in wb.worksheets:
        print("sheet", ws.title, "max", ws.max_row, ws.max_column, "dimension", ws.calculate_dimension())
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5), values_only=True):
            print(row[:30])
    wb.close()


def inspect_zip():
    with zipfile.ZipFile(FILE) as zf:
        print("zip entries worksheet/sharedStrings")
        for name in zf.namelist():
            if name.startswith("xl/worksheets/") or name in ("xl/sharedStrings.xml", "xl/workbook.xml"):
                info = zf.getinfo(name)
                print(name, info.file_size)
        for name in [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]:
            xml = zf.read(name).decode("utf-8", "ignore")
            print(name, "cell tags", xml.count("<c "), "row tags", xml.count("<row "))
            print(xml[:1000])


if __name__ == "__main__":
    print(FILE)
    inspect_openpyxl()
    inspect_zip()
