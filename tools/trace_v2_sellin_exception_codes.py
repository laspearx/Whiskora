import json
import openpyxl
from collections import defaultdict

V2 = r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Cover Day Report May 2026 - V.2.xlsx"
MONTHLY = r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Monthly - May 2026.xlsx"
SELLIN = r"C:\Work\Nestle\Target & Incentive\Sell in Query.xlsx"
MASTER = r"C:\Work\Nestle\Master\Final Master.xlsx"

TARGET_CODES = {6929156, 6903922, 6910470, 6904647, 7490540, 5525423, 200731, 5899801, 5011763}
MONTHS = {"03.2026", "04.2026", "05.2026"}

SELL_QTY_HEADERS = ["Sell in (1+2+3)/3", "(Last 3 Months)/3", "(Aug'25-Oct'25)/3"]
CODE_HEADERS = ["Nestlé Code", "Nestlรฉ Code", "Nestlเธฃเธ\x89 Code"]


def headers(ws):
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return h, {v: i for i, v in enumerate(h)}


def n(v):
    return v if isinstance(v, (int, float)) else 0


def code_col(idx):
    for col in CODE_HEADERS:
        if col in idx:
            return col
    raise KeyError("store code")


def read_v2():
    wb = openpyxl.load_workbook(V2, read_only=True, data_only=True)
    ws = wb["Raw Data"]
    h, idx = headers(ws)
    ccol = code_col(idx)
    by_code = defaultdict(lambda: {"rows": 0, "sell_qty": 0, "sell_baht": 0, "names": set(), "products": set()})
    by_name_hit = []
    target_names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if "Month" in idx and row[idx["Month"]] != "05_May-26":
            continue
        if row[idx["Channel"]] != "TT":
            continue
        code = row[idx[ccol]]
        name = row[idx["Store Name"]]
        product = row[idx["Product Name"]]
        sell_qty = sum(n(row[idx[col]]) for col in SELL_QTY_HEADERS if col in idx)
        sell_baht = n(row[idx[" Total Sell in (Baht) "]])
        if code in TARGET_CODES:
            b = by_code[code]
            b["rows"] += 1
            b["sell_qty"] += sell_qty
            b["sell_baht"] += sell_baht
            b["names"].add(name)
            b["products"].add(product)
    wb.close()
    return {
        str(code): {
            "rows": data["rows"],
            "sell_qty": data["sell_qty"],
            "sell_baht": data["sell_baht"],
            "names": sorted(data["names"]),
            "products": len(data["products"]),
        }
        for code, data in sorted(by_code.items())
    }


def read_monthly_exception_codes():
    wb = openpyxl.load_workbook(MONTHLY, read_only=True, data_only=True)
    ws = wb["Sell-in Mapping Exceptions"]
    h, idx = headers(ws)
    out = defaultdict(lambda: {"qty": 0, "rows": 0, "name": None, "products": set()})
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[idx["Customer Code"]]
        if code not in TARGET_CODES:
            continue
        b = out[code]
        b["name"] = row[idx["Customer Name"]]
        b["products"].add(row[idx["Product Name"]])
        b["rows"] += row[-2] or 0
        b["qty"] += row[-1] or 0
    wb.close()
    return {str(k): {"name": v["name"], "rows": v["rows"], "qty": v["qty"], "products": len(v["products"])} for k, v in sorted(out.items())}


def read_sellin_raw_codes():
    wb = openpyxl.load_workbook(SELLIN, read_only=True, data_only=True)
    ws = wb["Sell in Query"]
    h, idx = headers(ws)
    out = defaultdict(lambda: {"qty": 0, "nps": 0, "rows": 0, "name": None, "products": set(), "months": set()})
    for row in ws.iter_rows(min_row=2, values_only=True):
        month = str(row[idx["Cal. year / month"]] or "")
        code = row[idx["Customer"]]
        if month not in MONTHS or code not in TARGET_CODES:
            continue
        qty = max(n(row[idx["PCS."]]), 0)
        b = out[code]
        b["name"] = row[idx["Customer Name"]]
        b["rows"] += 1
        b["qty"] += qty
        b["nps"] += n(row[idx["NPS - THB"]])
        b["products"].add(row[idx["Product Name (CloudViu)"]])
        b["months"].add(month)
    wb.close()
    return {str(k): {"name": v["name"], "rows": v["rows"], "qty": v["qty"], "nps": v["nps"], "products": len(v["products"]), "months": sorted(v["months"])} for k, v in sorted(out.items())}


def read_master_hits():
    wb = openpyxl.load_workbook(MASTER, read_only=True, data_only=True)
    ws = wb["Store Master"]
    h = [c.value for c in next(ws.iter_rows(min_row=15, max_row=15))]
    idx = {v: i for i, v in enumerate(h)}
    hits = []
    for rn, row in enumerate(ws.iter_rows(min_row=16, values_only=True), start=16):
        vals = {col: row[i] if i < len(row) else None for col, i in idx.items()}
        codes = {vals.get("Store Code"), vals.get("storeCode"), vals.get("Group Code")}
        text = " ".join(str(v) for v in codes if v is not None)
        if any(str(code) in text for code in TARGET_CODES):
            hits.append({
                "row": rn,
                "Store Code": vals.get("Store Code"),
                "storeCode": vals.get("storeCode"),
                "Group Code": vals.get("Group Code"),
                "CloudViu": vals.get("Store Name (CloudViu)"),
                "SCVD": vals.get("Store Name (SCVD)"),
                "Channel": vals.get("Channel"),
                "Type": vals.get("Type"),
            })
    wb.close()
    return hits


def main():
    print(json.dumps({
        "sellin_raw_mar_apr_may_target_codes": read_sellin_raw_codes(),
        "monthly_exceptions_target_codes": read_monthly_exception_codes(),
        "v2_rows_with_same_codes": read_v2(),
        "final_master_hits": read_master_hits(),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
