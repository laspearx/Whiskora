from collections import defaultdict
from pathlib import Path
import re

import openpyxl


BASE = Path(r"C:\Work\Nestle")
REPORT_DIR = BASE / r"Report\(5th) Stock Cover Day"
MASTER_FILE = BASE / r"Master\Final Master.xlsx"
RAW_STOCK_FILE = REPORT_DIR / r"Raw Data Stock\Stock 31 May 2026.xlsx"
NEW_FILE = REPORT_DIR / "Stock Monthly - May 2026.xlsx"
V2_FILE = REPORT_DIR / "Stock Cover Day Report May 2026 - V.2.xlsx"


def as_int(value):
    if value is None:
        return None
    try:
        return int(float(str(value).strip()))
    except ValueError:
        return None


def norm(value):
    return str(value or "").replace(" (โฉมใหม่)", "").strip()


def number_parts(value):
    return [int(item) for item in re.findall(r"\d+", str(value or ""))]


def headers(ws, row=1):
    values = [c.value for c in next(ws.iter_rows(min_row=row, max_row=row))]
    return values, {h: i for i, h in enumerate(values)}


def load_store_master():
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws = wb["Store Master"]
    h, idx = headers(ws, 15)
    code_to_latest = {}
    latest_meta = {}
    code_meta = {}
    for row in ws.iter_rows(min_row=16, values_only=True):
        if all(v is None for v in row):
            continue
        code = as_int(row[idx["Store Code"]])
        if code is None:
            continue
        latest = max(number_parts(row[idx["Group Code"]]) or [code])
        meta = {
            "code": code,
            "latest": latest,
            "group": row[idx["Group Code"]],
            "channel": row[idx["Channel"]],
            "type": row[idx["Type"]],
            "name": row[idx["Store Name (SCVD)"]] or row[idx["Store Name (CloudViu)"]],
        }
        code_to_latest[code] = latest
        code_meta[code] = meta
        existing = latest_meta.get(latest)
        if existing is None or code == latest:
            latest_meta[latest] = meta
    wb.close()
    return code_to_latest, code_meta, latest_meta


def load_product_map():
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws = wb["Product Master"]
    h, idx = headers(ws, 14)
    exact = {}
    for row in ws.iter_rows(min_row=15, values_only=True):
        names = [row[idx.get("name")], row[idx.get("Product Name")], row[idx.get("Product Name (โฉมใหม่)")]]
        for name in names:
            if name:
                exact.setdefault(str(name).strip(), norm(name))
    wb.close()
    return exact


def raw_stock_by_latest():
    code_to_latest, code_meta, latest_meta = load_store_master()
    product_exact = load_product_map()
    wb = openpyxl.load_workbook(RAW_STOCK_FILE, read_only=False, data_only=True)
    ws = wb["Campaign 299"]
    h, idx = headers(ws)
    by_latest = defaultdict(float)
    by_channel = defaultdict(float)
    unmapped_store = defaultdict(float)
    raw_codes = defaultdict(float)
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = as_int(row[idx["store_code"]])
        qty = row[idx["value"]] or 0
        product = product_exact.get(str(row[idx["item_product_name"]] or "").strip(), norm(row[idx["item_product_name"]]))
        latest = code_to_latest.get(code, code)
        key = (latest, product)
        by_latest[key] += qty
        raw_codes[code] += qty
        meta = latest_meta.get(latest) or code_meta.get(code)
        ch = meta["channel"] if meta else "UNMAPPED"
        by_channel[ch] += qty
        if latest not in latest_meta:
            unmapped_store[(code, latest)] += qty
    wb.close()
    return by_latest, by_channel, unmapped_store, raw_codes, latest_meta


def output_stock(path, only_month=True):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Raw Data"]
    h, idx = headers(ws)
    by_key = defaultdict(float)
    by_channel = defaultdict(float)
    by_store = defaultdict(float)
    by_sell_key = defaultdict(float)
    by_sell_baht_key = defaultdict(float)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if only_month and idx.get("Month") is not None and row[idx["Month"]] != "05_May-26":
            continue
        code = as_int(row[idx["Nestlé Code"]])
        product = norm(row[idx["Product Name"]])
        qty = row[idx["#Total Stock"]] or 0
        sell_qty = (
            (row[idx.get("Sell in (1+2+3)/3")] or 0) if idx.get("Sell in (1+2+3)/3") is not None else 0
        ) + (
            (row[idx.get("(Last 3 Months)/3")] or 0) if idx.get("(Last 3 Months)/3") is not None else 0
        ) + (
            (row[idx.get("(Aug'25-Oct'25)/3")] or 0) if idx.get("(Aug'25-Oct'25)/3") is not None else 0
        )
        sell_baht = row[idx[" Total Sell in (Baht) "]] or 0
        by_key[(code, product)] += qty
        by_sell_key[(code, product)] += sell_qty
        by_sell_baht_key[(code, product)] += sell_baht
        by_channel[row[idx["Channel"]]] += qty
        by_store[(code, row[idx["Channel"]], row[idx["Store Name"]])] += qty
    wb.close()
    return by_key, by_channel, by_store, by_sell_key, by_sell_baht_key


def main():
    raw_key, raw_channel, unmapped, raw_codes, latest_meta = raw_stock_by_latest()
    new_key, new_channel, new_store, new_sell, new_sell_baht = output_stock(NEW_FILE)
    v2_key, v2_channel, v2_store, v2_sell, v2_sell_baht = output_stock(V2_FILE)

    print("RAW by channel from Store Master latest mapping", dict(raw_channel))
    print("NEW by channel", dict(new_channel))
    print("V2 by channel", dict(v2_channel))
    print("Unmapped raw stores", len(unmapped), sum(unmapped.values()))
    for item, qty in sorted(unmapped.items(), key=lambda kv: -kv[1])[:30]:
        print("UNMAPPED", item, qty)

    # Raw TT that does not land in new TT should reveal store/channel mapping issues.
    raw_tt = {k: v for k, v in raw_key.items() if (latest_meta.get(k[0]) or {}).get("channel") == "TT"}
    print("raw TT total", sum(raw_tt.values()), "keys", len(raw_tt))
    print("new TT total", new_channel.get("TT", 0), "v2 TT total", v2_channel.get("TT", 0))

    missing_vs_v2 = []
    for key, qty in v2_key.items():
        diff = qty - new_key.get(key, 0)
        if abs(diff) > 0.0001:
            missing_vs_v2.append((key, diff, qty, new_key.get(key, 0)))
    print("V2 minus NEW differing keys", len(missing_vs_v2), "total diff", sum(x[1] for x in missing_vs_v2))
    for key, diff, v2_qty, new_qty in sorted(missing_vs_v2, key=lambda x: -abs(x[1]))[:40]:
        print("DIFF", key, "diff", diff, "v2", v2_qty, "new", new_qty, "channel", (latest_meta.get(key[0]) or {}).get("channel"))

    print("Top V2 stores not in NEW or lower")
    store_diffs = []
    new_store_simple = defaultdict(float)
    for (code, _ch, _name), qty in new_store.items():
        new_store_simple[code] += qty
    for (code, ch, name), qty in v2_store.items():
        diff = qty - new_store_simple.get(code, 0)
        if abs(diff) > 0.0001:
            store_diffs.append((code, ch, name, diff, qty, new_store_simple.get(code, 0)))
    for item in sorted(store_diffs, key=lambda x: -abs(x[3]))[:40]:
        print("STORE_DIFF", item)

    sell_diffs = []
    for key, qty in v2_sell.items():
        diff = qty - new_sell.get(key, 0)
        if abs(diff) > 0.0001:
            sell_diffs.append((key, diff, qty, new_sell.get(key, 0), v2_sell_baht.get(key, 0), new_sell_baht.get(key, 0)))
    print("V2 minus NEW sell differing keys", len(sell_diffs), "total qty diff", sum(x[1] for x in sell_diffs), "total baht diff", sum(x[4] - x[5] for x in sell_diffs))
    for key, diff, v2_qty, new_qty, v2_baht, new_baht in sorted(sell_diffs, key=lambda x: -abs(x[1]))[:40]:
        print("SELL_DIFF", key, "qtydiff", diff, "v2", v2_qty, "new", new_qty, "bahtdiff", v2_baht - new_baht, "channel", (latest_meta.get(key[0]) or {}).get("channel"))


if __name__ == "__main__":
    main()
