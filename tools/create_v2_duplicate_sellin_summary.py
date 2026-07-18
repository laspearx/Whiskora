from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPORT_DIR = Path(r"C:\Work\Nestle\Report\(5th) Stock Cover Day")
V2_FILE = REPORT_DIR / "Stock Cover Day Report May 2026 - V.2.xlsx"
OUTPUT_FILE = REPORT_DIR / "V2 Duplicate Sell-in Summary May 2026.xlsx"

CODE_HEADERS = ["Nestlé Code", "Nestlรฉ Code", "Nestlเธฃเธ Code", "Nestlเน€เธเธเน€เธ\x89 Code"]
GROUP_CODE_HEADERS = ["Nestlé Group Code", "Nestlรฉ Group Code", "Nestlเธฃเธ Group Code"]
SELL_QTY_HEADERS = ["Sell in (1+2+3)/3", "(Last 3 Months)/3", "(Aug'25-Oct'25)/3"]
SELL_BAHT_HEADER = " Total Sell in (Baht) "
STOCK_BAHT_HEADER = "Total Stock (Baht)"
REASON_FIELDS = [
    "Area", "New Region", "P_Group", "Category", "P_Code", "Product SKU", "Size",
    "Brand", "Segment", "Group", "Nestlé Group Code", "Type", "Province", "outlet_id",
    "product_id", "Concat", "Concat2", " RSP LTP = TT (-1) ", "#Stock คงเหลือ โฉมปัจจุบัน",
    "#Stock คงเหลือ โฉมใหม่", "#Total Stock", "Total Stock (Baht)", "Sell in = TT (1)",
    "Sell in = TT (2)", "Sell in = TT (3)", "(Last 3 Months)/3", SELL_BAHT_HEADER,
]


def n(value):
    return value if isinstance(value, (int, float)) else 0


def headers(ws):
    row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return row, {value: index for index, value in enumerate(row)}


def first_existing(idx, names, fallback=None):
    for name in names:
        if name in idx:
            return name
    if fallback is not None:
        return fallback
    raise KeyError(names[0])


def row_value(row, idx, name, default=None):
    if name in idx:
        return row[idx[name]]
    return default


def collect_duplicates():
    wb = openpyxl.load_workbook(V2_FILE, read_only=True, data_only=True)
    ws = wb["Raw Data"]
    _header, idx = headers(ws)
    code_col = first_existing(idx, CODE_HEADERS)
    group_code_col = first_existing(idx, GROUP_CODE_HEADERS, None)

    groups = defaultdict(
        lambda: {
            "rows": 0,
            "positive_rows": 0,
            "source_rows": [],
            "row_sell_baht": [],
            "sell_qty": 0,
            "sell_baht": 0,
            "stock_baht": 0,
            "meta": None,
            "field_values": defaultdict(set),
        }
    )

    total_sell_baht = 0
    total_rows = 0
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if "Month" in idx and row[idx["Month"]] != "05_May-26":
            continue
        if row_value(row, idx, "Channel") != "TT":
            continue

        sell_baht = n(row_value(row, idx, SELL_BAHT_HEADER))
        sell_qty = sum(n(row_value(row, idx, col)) for col in SELL_QTY_HEADERS if col in idx)
        stock_baht = n(row_value(row, idx, STOCK_BAHT_HEADER))
        total_sell_baht += sell_baht
        total_rows += 1

        key = (
            row_value(row, idx, "Month", "05_May-26"),
            row_value(row, idx, code_col),
            row_value(row, idx, "Store Name"),
            row_value(row, idx, "Product Name"),
        )
        group = groups[key]
        group["rows"] += 1
        group["positive_rows"] += 1 if sell_baht else 0
        group["source_rows"].append(row_num)
        group["row_sell_baht"].append(sell_baht)
        group["sell_qty"] += sell_qty
        group["sell_baht"] += sell_baht
        group["stock_baht"] += stock_baht
        for field in REASON_FIELDS:
            if field in idx:
                group["field_values"][field].add(row_value(row, idx, field))
        if group["meta"] is None:
            group["meta"] = {
                "month": row_value(row, idx, "Month", "05_May-26"),
                "area": row_value(row, idx, "Area"),
                "region": row_value(row, idx, "New Region"),
                "store_code": row_value(row, idx, code_col),
                "group_code": row_value(row, idx, group_code_col) if group_code_col else None,
                "channel": row_value(row, idx, "Channel"),
                "type": row_value(row, idx, "Type"),
                "store_name": row_value(row, idx, "Store Name"),
                "province": row_value(row, idx, "Province"),
                "p_group": row_value(row, idx, "P_Group"),
                "category": row_value(row, idx, "Category"),
                "product_name": row_value(row, idx, "Product Name"),
                "product_sku": row_value(row, idx, "Product SKU"),
            }

    wb.close()

    detail = []
    for data in groups.values():
        if data["rows"] <= 1 or data["sell_baht"] <= 0:
            continue
        expected = max(data["row_sell_baht"])
        duplicated = data["sell_baht"] - expected
        if duplicated <= 0:
            continue
        reason, different_fields = classify_reason(data["field_values"])
        detail.append({**data["meta"], **{
            "reason": reason,
            "fields_different": ", ".join(different_fields),
            "duplicate_rows": data["rows"],
            "positive_sellin_rows": data["positive_rows"],
            "source_rows": ", ".join(str(row) for row in data["source_rows"]),
            "sell_qty_sum": data["sell_qty"],
            "v2_sell_baht_sum": data["sell_baht"],
            "expected_sell_baht": expected,
            "duplicated_sell_baht": duplicated,
            "stock_baht_sum": data["stock_baht"],
        }})

    detail.sort(key=lambda item: item["duplicated_sell_baht"], reverse=True)
    return detail, total_rows, total_sell_baht


def classify_reason(field_values):
    different_fields = [field for field, values in field_values.items() if len(values) > 1]
    if any(field in different_fields for field in ["Area", "New Region", "Type", "Province"]):
        return "Same store code/name/product appears under multiple master rows or regions", different_fields
    if any(field in different_fields for field in ["Concat", "Concat2", "outlet_id"]):
        return "Same store+product duplicated by helper key/outlet fields", different_fields
    return "Duplicate store+product rows with sell-in on more than one row", different_fields


def add_table(ws, headers_list, rows):
    ws.append(headers_list)
    for row in rows:
        ws.append(row)
    style_worksheet(ws)


def style_worksheet(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    for col in range(1, ws.max_column + 1):
        width = 14
        header = ws.cell(1, col).value
        if header in {"Store Name", "Product Name", "Source Rows"}:
            width = 42
        elif header in {"Explanation", "Note"}:
            width = 70
        elif header and "Baht" in str(header):
            width = 18
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
    ws.sheet_view.showGridLines = False


def summarize(detail, key_fields):
    summary = defaultdict(lambda: {"groups": 0, "rows": 0, "sell_sum": 0, "expected": 0, "duplicated": 0})
    for item in detail:
        key = tuple(item[field] for field in key_fields)
        target = summary[key]
        target["groups"] += 1
        target["rows"] += item["duplicate_rows"]
        target["sell_sum"] += item["v2_sell_baht_sum"]
        target["expected"] += item["expected_sell_baht"]
        target["duplicated"] += item["duplicated_sell_baht"]
    rows = []
    for key, value in summary.items():
        rows.append(list(key) + [value["groups"], value["rows"], value["sell_sum"], value["expected"], value["duplicated"]])
    rows.sort(key=lambda row: row[-1], reverse=True)
    return rows


def build_workbook():
    detail, total_rows, total_sell_baht = collect_duplicates()
    duplicated_total = sum(item["duplicated_sell_baht"] for item in detail)
    duplicate_group_sell_total = sum(item["v2_sell_baht_sum"] for item in detail)
    expected_total = sum(item["expected_sell_baht"] for item in detail)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    add_table(
        ws,
        ["Metric", "Value", "Explanation"],
        [
            ["Source file", str(V2_FILE), "May 2026 V.2 Raw Data, TT only"],
            ["Duplicate definition", "Same Month + Store Code + Store Name + Product Name", "If more than one row has sell-in for the same key, V.2 can count sell-in more than once."],
            ["V.2 TT rows checked", total_rows, "Rows in Raw Data for Month = 05_May-26 and Channel = TT"],
            ["Duplicate groups with sell-in", len(detail), "Store+SKU groups where duplicated sell-in baht is greater than zero"],
            ["V.2 sell-in baht inside duplicate groups", duplicate_group_sell_total, "Total sell-in baht before removing duplicate rows in the affected groups"],
            ["Expected sell-in baht inside duplicate groups", expected_total, "Conservative basis: keep the highest one row per duplicated store+SKU group"],
            ["Potential duplicated sell-in baht", duplicated_total, "Difference between V.2 group total and one retained row per group"],
            ["V.2 total TT sell-in baht checked", total_sell_baht, "Reference total from V.2 May TT raw data"],
        ],
    )

    detail_ws = wb.create_sheet("Duplicate Detail")
    detail_headers = [
        "Month", "Area", "New Region", "Store Code", "Group Code", "Channel", "Type",
        "Store Name", "Province", "P_Group", "Category", "Product Name", "Product SKU",
        "Reason", "Fields Different", "Duplicate Rows", "Positive Sell-in Rows", "Source Rows", "Sell-in Qty Sum",
        "V.2 Sell-in Baht Sum", "Expected Sell-in Baht", "Potential Duplicated Baht",
        "Stock Baht Sum",
    ]
    detail_rows = [
        [
            item["month"], item["area"], item["region"], item["store_code"], item["group_code"],
            item["channel"], item["type"], item["store_name"], item["province"], item["p_group"],
            item["category"], item["product_name"], item["product_sku"], item["reason"], item["fields_different"], item["duplicate_rows"],
            item["positive_sellin_rows"], item["source_rows"], item["sell_qty_sum"],
            item["v2_sell_baht_sum"], item["expected_sell_baht"], item["duplicated_sell_baht"],
            item["stock_baht_sum"],
        ]
        for item in detail
    ]
    add_table(detail_ws, detail_headers, detail_rows)

    reason_ws = wb.create_sheet("Reason Summary")
    reason_data = defaultdict(lambda: {"groups": 0, "rows": 0, "sell_sum": 0, "expected": 0, "duplicated": 0})
    for item in detail:
        target = reason_data[item["reason"]]
        target["groups"] += 1
        target["rows"] += item["duplicate_rows"]
        target["sell_sum"] += item["v2_sell_baht_sum"]
        target["expected"] += item["expected_sell_baht"]
        target["duplicated"] += item["duplicated_sell_baht"]
    add_table(
        reason_ws,
        ["Reason", "Duplicate Groups", "Duplicate Rows", "V.2 Sell-in Baht Sum", "Expected Sell-in Baht", "Potential Duplicated Baht", "Customer Explanation"],
        [
            [
                reason,
                data["groups"],
                data["rows"],
                data["sell_sum"],
                data["expected"],
                data["duplicated"],
                "Sell-in was mapped to more than one raw row for the same store code, store name, and product; the corrected file keeps the sale once.",
            ]
            for reason, data in sorted(reason_data.items(), key=lambda item: item[1]["duplicated"], reverse=True)
        ],
    )

    store_ws = wb.create_sheet("Summary by Store")
    add_table(
        store_ws,
        ["Store Code", "Store Name", "Type", "Duplicate Groups", "Duplicate Rows", "V.2 Sell-in Baht Sum", "Expected Sell-in Baht", "Potential Duplicated Baht"],
        summarize(detail, ["store_code", "store_name", "type"]),
    )

    product_ws = wb.create_sheet("Summary by Product")
    add_table(
        product_ws,
        ["P_Group", "Category", "Product Name", "Product SKU", "Duplicate Groups", "Duplicate Rows", "V.2 Sell-in Baht Sum", "Expected Sell-in Baht", "Potential Duplicated Baht"],
        summarize(detail, ["p_group", "category", "product_name", "product_sku"]),
    )

    pgroup_ws = wb.create_sheet("Summary by P_Group")
    add_table(
        pgroup_ws,
        ["P_Group", "Category", "Duplicate Groups", "Duplicate Rows", "V.2 Sell-in Baht Sum", "Expected Sell-in Baht", "Potential Duplicated Baht"],
        summarize(detail, ["p_group", "category"]),
    )

    wb.save(OUTPUT_FILE)
    return {
        "output": str(OUTPUT_FILE),
        "duplicate_groups": len(detail),
        "duplicated_total": duplicated_total,
        "duplicate_group_sell_total": duplicate_group_sell_total,
        "expected_total": expected_total,
        "top_5": detail[:5],
    }


if __name__ == "__main__":
    print(build_workbook())
