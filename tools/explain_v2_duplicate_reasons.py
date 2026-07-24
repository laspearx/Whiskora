from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


REPORT_DIR = Path(r"C:\Work\Nestle\Report\(5th) Stock Cover Day")
V2_FILE = REPORT_DIR / "Stock Cover Day Report May 2026 - V.2.xlsx"


CODE_HEADERS = ["Nestlé Code", "Nestlรฉ Code", "Nestlเธฃเธ Code"]
SELL_BAHT_HEADER = " Total Sell in (Baht) "


def n(value):
    return value if isinstance(value, (int, float)) else 0


def headers(ws):
    row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return row, {value: index for index, value in enumerate(row)}


def first_existing(idx, names):
    for name in names:
        if name in idx:
            return name
    raise KeyError(names[0])


def row_dict(header, values):
    return {header[index]: value for index, value in enumerate(values) if index < len(header)}


def reason_for(rows):
    varying = []
    for field in [
        "Area",
        "New Region",
        "P_Group",
        "Category",
        "P_Code",
        "Product SKU",
        "Size",
        "Brand",
        "Segment",
        "Group",
        "Nestlé Group Code",
        "Type",
        "Province",
        "outlet_id",
        "product_id",
        "Concat",
        "Concat2",
        " RSP LTP = TT (-1) ",
        "#Stock คงเหลือ โฉมปัจจุบัน",
        "#Stock คงเหลือ โฉมใหม่",
        "#Total Stock",
        "Total Stock (Baht)",
        "Sell in = TT (1)",
        "Sell in = TT (2)",
        "Sell in = TT (3)",
        "(Last 3 Months)/3",
        SELL_BAHT_HEADER,
    ]:
        values = {row.get(field) for row in rows}
        if len(values) > 1:
            varying.append(field)

    if any(field in varying for field in ["Area", "New Region", "Type", "Province"]):
        return "same store code/name/product appears under multiple master rows or regions"
    if any(field in varying for field in ["Concat", "Concat2", "outlet_id"]):
        return "same store+product duplicated by helper key/outlet fields"
    if all(row.get(SELL_BAHT_HEADER) for row in rows) and len({row.get(SELL_BAHT_HEADER) for row in rows}) == 1:
        return "identical sell-in copied to duplicate rows"
    return "duplicate store+product rows with sell-in on more than one row"


def main():
    wb = openpyxl.load_workbook(V2_FILE, read_only=True, data_only=True)
    ws = wb["Raw Data"]
    header, idx = headers(ws)
    code_col = first_existing(idx, CODE_HEADERS)

    groups = defaultdict(list)
    for row_num, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = row_dict(header, values)
        if data.get("Month") != "05_May-26" or data.get("Channel") != "TT":
            continue
        if not n(data.get(SELL_BAHT_HEADER)):
            continue
        key = (data.get("Month"), data.get(code_col), data.get("Store Name"), data.get("Product Name"))
        data["_row_num"] = row_num
        groups[key].append(data)
    wb.close()

    reason_counts = Counter()
    reason_baht = Counter()
    store_code_counts = Counter()
    duplicate_details = []
    for key, rows in groups.items():
        if len(rows) <= 1:
            continue
        sell_values = [n(row.get(SELL_BAHT_HEADER)) for row in rows]
        duplicated_baht = sum(sell_values) - max(sell_values)
        if duplicated_baht <= 0:
            continue
        reason = reason_for(rows)
        reason_counts[reason] += 1
        reason_baht[reason] += duplicated_baht
        store_code_counts[key[1]] += 1
        duplicate_details.append((duplicated_baht, reason, key, [row["_row_num"] for row in rows], rows))

    print("Reason summary")
    for reason, count in reason_counts.most_common():
        print(f"{count}\t{reason}\t{reason_baht[reason]:,.2f}")

    print("\nTop repeated store codes")
    for code, count in store_code_counts.most_common(15):
        print(f"{code}\t{count}")

    print("\nTop examples")
    for duplicated_baht, reason, key, row_nums, rows in sorted(duplicate_details, reverse=True)[:10]:
        compared = {
            "key": key,
            "rows": row_nums,
            "reason": reason,
            "duplicated_baht": round(duplicated_baht, 2),
            "areas": sorted({str(row.get("Area")) for row in rows}),
            "regions": sorted({str(row.get("New Region")) for row in rows}),
            "types": sorted({str(row.get("Type")) for row in rows}),
            "provinces": sorted({str(row.get("Province")) for row in rows}),
            "group_codes": sorted({str(row.get("Nestlé Group Code")) for row in rows}),
            "sell_baht_rows": [round(n(row.get(SELL_BAHT_HEADER)), 2) for row in rows],
            "stock_baht_rows": [round(n(row.get("Total Stock (Baht)")), 2) for row in rows],
        }
        print(compared)


if __name__ == "__main__":
    main()
