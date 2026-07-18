import json
import openpyxl
from collections import defaultdict

V2 = r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Cover Day Report May 2026 - V.2.xlsx"
MONTHLY = r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Monthly - May 2026.xlsx"

CODE_HEADERS = ["Nestlé Code", "Nestlรฉ Code", "Nestlเธฃเธ\x89 Code"]
SELL_QTY_HEADERS = ["Sell in (1+2+3)/3", "(Last 3 Months)/3", "(Aug'25-Oct'25)/3"]


def headers(ws):
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return h, {v: i for i, v in enumerate(h)}


def code_col(idx):
    for col in CODE_HEADERS:
        if col in idx:
            return col
    raise KeyError("code")


def n(v):
    return v if isinstance(v, (int, float)) else 0


def duplicates(path, label):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Raw Data"]
    h, idx = headers(ws)
    ccol = code_col(idx)
    groups = defaultdict(lambda: {"rows": 0, "sell_qty": 0, "sell_baht": 0, "stock_baht": 0, "examples": [], "row_sell_baht": []})
    for rn, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if label == "v2" and "Month" in idx and row[idx["Month"]] != "05_May-26":
            continue
        if row[idx["Channel"]] != "TT":
            continue
        key = (row[idx[ccol]], row[idx["Store Name"]], row[idx["Product Name"]])
        g = groups[key]
        g["rows"] += 1
        g["sell_qty"] += sum(n(row[idx[col]]) for col in SELL_QTY_HEADERS if col in idx)
        g["sell_baht"] += n(row[idx[" Total Sell in (Baht) "]])
        g["stock_baht"] += n(row[idx["Total Stock (Baht)"]])
        if len(g["examples"]) < 5:
            g["examples"].append(rn)
        g["row_sell_baht"].append(n(row[idx[" Total Sell in (Baht) "]]))
    wb.close()
    dup = []
    for (code, store, product), g in groups.items():
        if g["rows"] > 1:
            # Conservative potential overage: all rows beyond the first, assuming duplicated rows carry the same sell-in basis.
            over_baht = g["sell_baht"] - max(g["row_sell_baht"])
            dup.append([code, store, product, g["rows"], g["sell_qty"], g["sell_baht"], over_baht, g["stock_baht"], g["examples"]])
    return sorted(dup, key=lambda row: row[5], reverse=True)


def main():
    v2 = duplicates(V2, "v2")
    monthly = duplicates(MONTHLY, "monthly")
    print(json.dumps({
        "v2_duplicate_group_count": len(v2),
        "v2_duplicate_group_sell_baht_total": sum(row[5] for row in v2),
        "v2_duplicate_over_baht_total": sum(row[6] for row in v2),
        "v2_top_duplicates": v2[:30],
        "monthly_duplicate_group_count": len(monthly),
        "monthly_duplicate_group_sell_baht_total": sum(row[5] for row in monthly),
        "monthly_duplicate_over_baht_total": sum(row[6] for row in monthly),
        "monthly_top_duplicates": monthly[:10],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
