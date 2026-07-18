import collections
import json

import openpyxl


FILES = {
    "monthly": r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Monthly - May 2026.xlsx",
    "v2": r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Cover Day Report May 2026 - V.2.xlsx",
    "v3": r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Cover Day Report May 2026 - V.3.xlsx",
}


STOCK_CODE_HEADERS = ["Nestlé Code", "Nestlรฉ Code", "Nestlเธฃเธ\x89 Code"]
SELL_QTY_HEADERS = ["Sell in (1+2+3)/3", "(Last 3 Months)/3", "(Aug'25-Oct'25)/3"]


def header_map(ws):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return headers, {header: idx for idx, header in enumerate(headers)}


def num(row, idx, name):
    col = idx.get(name)
    if col is None:
        return 0
    value = row[col]
    return value if isinstance(value, (int, float)) else 0


def first_existing(idx, names):
    for name in names:
        if name in idx:
            return name
    return None


def summarize_raw(path, label):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Raw Data"]
    headers, idx = header_map(ws)
    code_header = first_existing(idx, STOCK_CODE_HEADERS)
    by_channel = collections.defaultdict(lambda: {"rows": 0, "stores": set(), "skus": set(), "stock_qty": 0, "stock_baht": 0, "sell_qty": 0, "sell_baht": 0})
    by_type = collections.defaultdict(lambda: {"rows": 0, "stores": set(), "skus": set(), "stock_qty": 0, "stock_baht": 0, "sell_qty": 0, "sell_baht": 0})
    store_totals = collections.defaultdict(lambda: {"stock_qty": 0, "stock_baht": 0, "sell_qty": 0, "sell_baht": 0, "rows": 0, "channel": None, "type": None, "name": None})

    for row in ws.iter_rows(min_row=2, values_only=True):
        if "Month" in idx and label in {"v2", "v3"} and row[idx["Month"]] != "05_May-26":
            continue
        ch = row[idx["Channel"]]
        typ = row[idx["Type"]]
        code = row[idx[code_header]] if code_header else None
        product = row[idx["Product Name"]]
        store_name = row[idx["Store Name"]]
        stock_qty = num(row, idx, "#Total Stock")
        stock_baht = num(row, idx, "Total Stock (Baht)")
        sell_qty = sum(num(row, idx, name) for name in SELL_QTY_HEADERS)
        sell_baht = num(row, idx, " Total Sell in (Baht) ")

        for key, bucket in [(ch, by_channel), (typ, by_type)]:
            target = bucket[key]
            target["rows"] += 1
            target["stores"].add((code, store_name))
            target["skus"].add(product)
            target["stock_qty"] += stock_qty
            target["stock_baht"] += stock_baht
            target["sell_qty"] += sell_qty
            target["sell_baht"] += sell_baht

        st = store_totals[(code, store_name)]
        st["rows"] += 1
        st["channel"] = ch
        st["type"] = typ
        st["name"] = store_name
        st["stock_qty"] += stock_qty
        st["stock_baht"] += stock_baht
        st["sell_qty"] += sell_qty
        st["sell_baht"] += sell_baht

    wb.close()
    return {
        "by_channel": compact(by_channel),
        "by_type": compact(by_type),
        "store_totals": store_totals,
    }


def compact(mapping):
    return {
        str(key): {
            name: (len(value) if isinstance(value, set) else value)
            for name, value in data.items()
        }
        for key, data in sorted(mapping.items(), key=lambda item: str(item[0]))
    }


def sheet_top(path, sheet_name, max_rows=10):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[sheet_name]
    rows = [list(row) for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=True)]
    wb.close()
    return rows


def exceptions(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for sheet in ["Summary", "Stock Mapping Exceptions", "Sell-in Mapping Exceptions"]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        if sheet == "Summary":
            out[sheet] = {str(r[0]): r[1] for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None}
        else:
            qty = 0
            rows = 0
            for r in ws.iter_rows(min_row=2, values_only=True):
                rows += r[-2] or 0
                qty += r[-1] or 0
            out[sheet] = {"source_rows": rows, "qty": qty, "exception_lines": ws.max_row - 1}
    wb.close()
    return out


def store_gap(v2_totals, monthly_totals, metric):
    rows = []
    keys = set(v2_totals) | set(monthly_totals)
    for key in keys:
        old = v2_totals.get(key, {}).get(metric, 0)
        new = monthly_totals.get(key, {}).get(metric, 0)
        diff = new - old
        if diff:
            rows.append([key[0], key[1], old, new, diff])
    return sorted(rows, key=lambda row: abs(row[4]), reverse=True)[:20]


def main():
    data = {label: summarize_raw(path, label) for label, path in FILES.items()}
    result = {
        "monthly_pivot_top": sheet_top(FILES["monthly"], "Pivot", 6),
        "v2_first_sheet": sheet_top(FILES["v2"], openpyxl.load_workbook(FILES["v2"], read_only=True).sheetnames[0], 8),
        "summaries": {
            label: {"by_channel": data[label]["by_channel"], "by_type": data[label]["by_type"]}
            for label in data
        },
        "monthly_exceptions": exceptions(FILES["monthly"]),
        "top_store_stock_baht_gap_v2_minus_monthly": store_gap(data["v2"]["store_totals"], data["monthly"]["store_totals"], "stock_baht"),
        "top_store_sell_baht_gap_v2_minus_monthly": store_gap(data["v2"]["store_totals"], data["monthly"]["store_totals"], "sell_baht"),
    }
    print(json.dumps(result, ensure_ascii=False, default=str, indent=2))


if __name__ == "__main__":
    main()
