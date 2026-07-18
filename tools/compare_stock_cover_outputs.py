import collections
import openpyxl


FILES = [
    ("new", r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Monthly - May 2026.xlsx"),
    ("v2", r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Cover Day Report May 2026 - V.2.xlsx"),
    ("orig", r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Cover Day Report May 2026.xlsx"),
]


def header_map(ws):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return headers, {header: idx for idx, header in enumerate(headers)}


def num(row, idx, name):
    col = idx.get(name)
    if col is None:
        return 0
    value = row[col]
    return value if isinstance(value, (int, float)) else 0


def main():
    for label, path in FILES:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Raw Data"]
        headers, idx = header_map(ws)
        by_channel = collections.defaultdict(lambda: [0, 0, 0, 0, 0, set(), set()])
        by_type = collections.defaultdict(lambda: [0, 0, 0, 0, 0, set(), set()])
        for row in ws.iter_rows(min_row=2, values_only=True):
            if idx.get("Month") is not None and row[idx["Month"]] != "05_May-26":
                continue
            sell_qty = (
                num(row, idx, "Sell in (1+2+3)/3")
                + num(row, idx, "(Last 3 Months)/3")
                + num(row, idx, "(Aug'25-Oct'25)/3")
            )
            metrics = [
                1,
                num(row, idx, "#Total Stock"),
                num(row, idx, "Total Stock (Baht)"),
                num(row, idx, " Total Sell in (Baht) "),
                sell_qty,
            ]
            for key, bucket in [
                (row[idx.get("Channel")], by_channel),
                (row[idx.get("Type")], by_type),
            ]:
                vals = bucket[key]
                for i, value in enumerate(metrics):
                    vals[i] += value
                vals[5].add(row[idx["Nestlé Code"]])
                vals[6].add(row[idx["Product Name"]])

        print(f"\n## {label}")
        print("By Channel")
        for key, vals in sorted(by_channel.items(), key=lambda item: str(item[0])):
            print(key, "rows", vals[0], "stores", len(vals[5]), "skus", len(vals[6]), "stockqty", vals[1], "stockbaht", vals[2], "sellbaht", vals[3], "sellqty", vals[4])
        print("By Type")
        for key, vals in sorted(by_type.items(), key=lambda item: str(item[0])):
            print(key, "rows", vals[0], "stores", len(vals[5]), "skus", len(vals[6]), "stockqty", vals[1], "stockbaht", vals[2], "sellbaht", vals[3], "sellqty", vals[4])
        wb.close()


if __name__ == "__main__":
    main()
