from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPORT_DIR = Path(r"C:\Work\Nestle\Report\(5th) Stock Cover Day")
MONTHLY_FILE = REPORT_DIR / "Stock Monthly - May 2026.xlsx"
REPORT_OUTPUT = REPORT_DIR / "Stock Cover Day Report May 2026 - V.3.xlsx"


def load_tt_rows():
    wb = openpyxl.load_workbook(MONTHLY_FILE, read_only=True, data_only=True)
    ws = wb["Raw Data"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: i for i, header in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Channel"]] == "TT":
            rows.append(list(row))
    wb.close()
    return headers, rows


def cvd(stock_baht, sell_baht):
    return stock_baht / sell_baht * 30 if sell_baht else 0


def style_sheet(ws, header_fill="1F4E78"):
    fill = PatternFill("solid", fgColor=header_fill)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False


def set_widths(ws, widths=None, default=14):
    widths = widths or {}
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, default)


def write_sheet(wb, name, headers, data_rows, widths=None):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in data_rows:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, widths)
    return ws


def aggregate(rows, idx):
    channel = defaultdict(lambda: {"rows": 0, "stores": set(), "skus": set(), "stock": 0, "stock_baht": 0, "sell": 0, "sell_baht": 0})
    sub_brand = defaultdict(lambda: {"rows": 0, "stores": set(), "skus": set(), "stock": 0, "stock_baht": 0, "sell": 0, "sell_baht": 0})
    store = defaultdict(lambda: {"meta": None, "skus": set(), "stock": 0, "stock_baht": 0, "sell": 0, "sell_baht": 0})
    store_sku = defaultdict(lambda: {"meta": None, "stock": 0, "stock_baht": 0, "sell": 0, "sell_baht": 0})

    for row in rows:
        stock = row[idx["#Total Stock"]] or 0
        stock_baht = row[idx["Total Stock (Baht)"]] or 0
        sell = row[idx["Sell in (1+2+3)/3"]] or 0
        sell_baht = row[idx[" Total Sell in (Baht) "]] or 0
        store_code = row[idx["Nestlé Code"]]
        product = row[idx["Product Name"]]

        ch_key = row[idx["Channel"]]
        target = channel[ch_key]
        target["rows"] += 1
        target["stores"].add(store_code)
        target["skus"].add(product)
        target["stock"] += stock
        target["stock_baht"] += stock_baht
        target["sell"] += sell
        target["sell_baht"] += sell_baht

        sb_key = (row[idx["P_Group"]], row[idx["Category"]])
        target = sub_brand[sb_key]
        target["rows"] += 1
        target["stores"].add(store_code)
        target["skus"].add(product)
        target["stock"] += stock
        target["stock_baht"] += stock_baht
        target["sell"] += sell
        target["sell_baht"] += sell_baht

        st_key = (store_code, row[idx["Store Name"]])
        target = store[st_key]
        target["meta"] = [
            row[idx["Area"]],
            row[idx["New Region"]],
            store_code,
            row[idx["Nestlé Group Code"]],
            row[idx["Channel"]],
            row[idx["Type"]],
            row[idx["Store Name"]],
            row[idx["Province"]],
        ]
        target["skus"].add(product)
        target["stock"] += stock
        target["stock_baht"] += stock_baht
        target["sell"] += sell
        target["sell_baht"] += sell_baht

        ss_key = (store_code, row[idx["Store Name"]], product)
        target = store_sku[ss_key]
        target["meta"] = [
            row[idx["Area"]],
            row[idx["New Region"]],
            store_code,
            row[idx["Nestlé Group Code"]],
            row[idx["Channel"]],
            row[idx["Type"]],
            row[idx["Store Name"]],
            row[idx["Province"]],
            row[idx["P_Group"]],
            row[idx["Category"]],
            row[idx["Product Name"]],
            row[idx["Product SKU"]],
        ]
        target["stock"] += stock
        target["stock_baht"] += stock_baht
        target["sell"] += sell
        target["sell_baht"] += sell_baht

    return channel, sub_brand, store, store_sku


def build_report():
    raw_headers, rows = load_tt_rows()
    idx = {header: i for i, header in enumerate(raw_headers)}
    channel, sub_brand, store, store_sku = aggregate(rows, idx)

    wb = Workbook()
    wb.remove(wb.active)

    channel_rows = []
    for key, data in sorted(channel.items(), key=lambda item: str(item[0])):
        channel_rows.append([
            key,
            data["rows"],
            len(data["stores"]),
            len(data["skus"]),
            data["stock"],
            data["stock_baht"],
            data["sell"],
            data["sell_baht"],
            cvd(data["stock_baht"], data["sell_baht"]),
        ])
    write_sheet(
        wb,
        "Summary By Channel",
        ["Channel", "Rows", "Stores", "SKUs", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"],
        channel_rows,
    )

    sub_brand_rows = []
    for (p_group, category), data in sorted(sub_brand.items(), key=lambda item: (str(item[0][0]), str(item[0][1]))):
        sub_brand_rows.append([
            p_group,
            category,
            data["rows"],
            len(data["stores"]),
            len(data["skus"]),
            data["stock"],
            data["stock_baht"],
            data["sell"],
            data["sell_baht"],
            cvd(data["stock_baht"], data["sell_baht"]),
        ])
    write_sheet(
        wb,
        "Summary By Sub Brand",
        ["P_Group", "Category", "Rows", "Stores", "SKUs", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"],
        sub_brand_rows,
        {1: 24, 2: 24},
    )

    stock_rows = []
    for (_code, _name), data in sorted(store.items(), key=lambda item: (str(item[0][0]), str(item[0][1]))):
        stock_rows.append(data["meta"] + [
            len(data["skus"]),
            data["stock"],
            data["stock_baht"],
            data["sell"],
            data["sell_baht"],
            cvd(data["stock_baht"], data["sell_baht"]),
        ])
    write_sheet(
        wb,
        "Stock Data",
        ["Area", "New Region", "Nestlé Code", "Nestlé Group Code", "Channel", "Type", "Store Name", "Province", "SKUs", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"],
        stock_rows,
        {7: 34},
    )

    write_sheet(
        wb,
        "Raw Data",
        raw_headers,
        rows,
        {3: 24, 4: 24, 6: 28, 7: 42, 16: 34, 21: 42, 22: 42},
    )

    pivot_store_rows = []
    for (_code, _name), data in sorted(store.items(), key=lambda item: (str(item[0][0]), str(item[0][1]))):
        pivot_store_rows.append([
            data["meta"][2],
            data["meta"][3],
            data["meta"][6],
            data["meta"][4],
            data["meta"][5],
            data["stock"],
            data["stock_baht"],
            data["sell"],
            data["sell_baht"],
            cvd(data["stock_baht"], data["sell_baht"]),
        ])
    write_sheet(
        wb,
        "Pivot by Store",
        ["Nestlé Code", "Nestlé Group Code", "Store Name", "Channel", "Type", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"],
        pivot_store_rows,
        {3: 34},
    )

    store_sku_rows = []
    for (_code, _name, _product), data in sorted(store_sku.items(), key=lambda item: (str(item[0][0]), str(item[0][1]), str(item[0][2]))):
        store_sku_rows.append(data["meta"] + [
            data["stock"],
            data["stock_baht"],
            data["sell"],
            data["sell_baht"],
            cvd(data["stock_baht"], data["sell_baht"]),
        ])
    write_sheet(
        wb,
        "Store by SKU",
        ["Area", "New Region", "Nestlé Code", "Nestlé Group Code", "Channel", "Type", "Store Name", "Province", "P_Group", "Category", "Product Name", "Product SKU", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"],
        store_sku_rows,
        {7: 34, 11: 42, 12: 28},
    )

    wb.save(REPORT_OUTPUT)
    return {
        "rows": len(rows),
        "stores": len(store),
        "skus": len({row[idx["Product Name"]] for row in rows}),
        "stock": sum(row[idx["#Total Stock"]] or 0 for row in rows),
        "stock_baht": sum(row[idx["Total Stock (Baht)"]] or 0 for row in rows),
        "sell": sum(row[idx["Sell in (1+2+3)/3"]] or 0 for row in rows),
        "sell_baht": sum(row[idx[" Total Sell in (Baht) "]] or 0 for row in rows),
    }


if __name__ == "__main__":
    print(build_report())
