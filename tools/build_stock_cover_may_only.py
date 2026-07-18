from collections import defaultdict
from datetime import datetime
from pathlib import Path
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\Work\Nestle")
REPORT_DIR = BASE / r"Report\(5th) Stock Cover Day"
MASTER_FILE = BASE / r"Master\Final Master.xlsx"
RSP_FILE = BASE / r"Master\RSP on month.xlsx"
RAW_STOCK_FILE = REPORT_DIR / r"Raw Data Stock\Stock 31 May 2026.xlsx"
SELL_IN_STAGING_FILE = REPORT_DIR / "Stock Cover Day.xlsx"
SELL_IN_QUERY_FILE = BASE / r"Target & Incentive\Sell in Query.xlsx"
MONTHLY_OUTPUT = REPORT_DIR / "Stock Monthly - May 2026.xlsx"

MONTH_LABEL = "05_May-26"
SELL_IN_MONTHS = ["03.2026", "04.2026", "05.2026"]
REPORT_DATE = datetime(2026, 6, 1)
YEAR = 2026

HEADERS = [
    "Area",
    "New Region",
    "P_Group",
    "Category",
    "P_Code",
    "Product SKU",
    "Product Name",
    "Size",
    "Brand",
    "Segment",
    "Group",
    "Nestlé Code",
    "Nestlé Group Code",
    "Channel",
    "Type",
    "Store Name",
    "Province",
    "Date",
    "outlet_id",
    "product_id",
    "Concat",
    "Concat2",
    " RSP LTP = TT (-1) ",
    "#Stock คงเหลือ โฉมปัจจุบัน",
    "#Stock คงเหลือ โฉมใหม่",
    "#Total Stock",
    "Total Stock (Baht)",
    "Sell in = TT (1)",
    "Sell in = TT (2)",
    "Sell in = TT (3)",
    "Sell in (1+2+3)/3",
    " Total Sell in (Baht) ",
    "Month",
    "Year",
]


def as_int(value):
    if value is None:
        return None
    text = str(value).strip()
    try:
        return int(float(text))
    except ValueError:
        match = re.search(r"\d+", text)
        return int(match.group(0)) if match else None


def norm_name(value):
    return str(value or "").replace(" (โฉมใหม่)", "").strip()


def match_key(value):
    return re.sub(r"\s+", "", str(value or "").strip().casefold())


def header_row(ws, row_num):
    return [cell.value for cell in next(ws.iter_rows(min_row=row_num, max_row=row_num))]


def dict_rows(ws, header_row_num, first_data_row):
    headers = header_row(ws, header_row_num)
    for values in ws.iter_rows(min_row=first_data_row, values_only=True):
        if all(value is None for value in values):
            continue
        yield {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}


def add_store_instance(stores, by_code, canonical, row, code, preferred_name=None):
    names = [
        preferred_name,
        row.get("Store Name (CloudViu)"),
        row.get("Store Name (SCVD)"),
        row.get("Store Name (Group)"),
        row.get("Store Name (TH)"),
        row.get("Store Name (EN)"),
    ]
    name_key = next((match_key(name) for name in names if match_key(name)), f"store-{code}")
    key = (code, name_key)
    if key not in stores:
        row = dict(row)
        row["_store_key"] = key
        row["_store_code"] = code
        row["_match_names"] = {match_key(name) for name in names if match_key(name)}
        stores[key] = row
        by_code[code].append(key)
        canonical.setdefault(code, key)
    return key


def index_store_alias(by_code, canonical, alias_code, key):
    alias_code = as_int(alias_code)
    if alias_code is None:
        return
    if key not in by_code[alias_code]:
        by_code[alias_code].append(key)
    canonical.setdefault(alias_code, key)


def load_store_master():
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws = wb["Store Master"]
    stores = {}
    by_code = defaultdict(list)
    canonical = {}
    for row in dict_rows(ws, 15, 16):
        code = as_int(row.get("Store Code"))
        if code is not None:
            key = add_store_instance(stores, by_code, canonical, row, code)
            index_store_alias(by_code, canonical, row.get("storeCode"), key)
    wb.close()
    return stores, by_code, canonical


def load_product_master():
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws = wb["Product Master"]
    by_exact = {}
    by_norm = {}
    for row in dict_rows(ws, 14, 15):
        exact_names = {
            str(row.get("name") or "").strip(),
            str(row.get("Product Name") or "").strip(),
            str(row.get("Product Name (โฉมใหม่)") or "").strip(),
        }
        for exact in exact_names:
            if exact:
                by_exact.setdefault(exact, row)
        normalized = norm_name(row.get("Product Name (โฉมใหม่)") or row.get("Product Name") or row.get("name"))
        if normalized:
            by_norm.setdefault(normalized, row)
    wb.close()
    return by_exact, by_norm


def load_rsp():
    wb = openpyxl.load_workbook(RSP_FILE, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = header_row(ws, 1)
    idx = {h: i for i, h in enumerate(headers)}
    rsp = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Month"]] == MONTH_LABEL:
            rsp[norm_name(row[idx["Product Name"]])] = row[idx["RSP"]] or 0
    wb.close()
    return rsp


def resolve_raw_store(raw_code, outlet_name, stores, by_code):
    raw_name_key = match_key(outlet_name)
    candidates = by_code.get(raw_code, [])
    for key in candidates:
        if raw_name_key and raw_name_key in stores[key].get("_match_names", set()):
            return key
    if len(candidates) == 1:
        return candidates[0]
    return None


def load_raw_stock(stores, by_code, canonical, product_by_exact, product_by_norm, rsp_by_name):
    # Normal mode is intentional: this workbook exposes only column A to
    # openpyxl read_only mode, while Excel and normal mode see all 24 columns.
    wb = openpyxl.load_workbook(RAW_STOCK_FILE, read_only=False, data_only=True)
    ws = wb["Campaign 299"]
    headers = header_row(ws, 1)
    idx = {h: i for i, h in enumerate(headers)}
    products = {}
    stock = defaultdict(lambda: {"current": 0, "new": 0})
    raw_store_keys = set()
    exceptions = defaultdict(lambda: {"rows": 0, "qty": 0})

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_code = as_int(row[idx["store_code"]])
        if raw_code is None:
            continue
        raw_name = str(row[idx["item_product_name"]] or "").strip()
        report_name = norm_name(raw_name)
        if not report_name:
            continue

        master = product_by_exact.get(raw_name) or product_by_norm.get(report_name) or {}
        variant = str(master.get("SPD") or "").strip()
        bucket = "new" if "ใหม่" in raw_name or "ใหม่" in variant else "current"
        rsp = rsp_by_name.get(report_name) or master.get("marketPrice") or 0
        products.setdefault(
            report_name,
            {
                "P_Group": master.get("P_group"),
                "Category": master.get("Category"),
                "P_Code": master.get("P_code"),
                "Product SKU": master.get("Attribute Group") or master.get("label") or row[idx["item_label"]],
                "Product Name": report_name,
                "Size": master.get("Size"),
                "Brand": master.get("Brand"),
                "Segment": master.get("Group"),
                "Group": master.get("Type"),
                "product_id": master.get("sku"),
                "rsp": rsp,
            },
        )
        qty = row[idx["value"]] or 0
        store_key = resolve_raw_store(raw_code, row[idx["outlet_name"]], stores, by_code)
        if store_key is None:
            reason = "Store code not in Final Master" if raw_code not in by_code else "Multiple Final Master stores share code and outlet name did not match"
            key = (raw_code, row[idx["outlet_id"]], row[idx["outlet_name"]], report_name, reason)
            exceptions[key]["rows"] += 1
            exceptions[key]["qty"] += qty
            continue
        stock[(store_key, report_name)][bucket] += qty
        raw_store_keys.add(store_key)

    wb.close()
    return products, stock, raw_store_keys, exceptions


def load_sell_in(stores, by_code, canonical, products):
    product_names = {norm_name(name): name for name in products}
    wb = openpyxl.load_workbook(SELL_IN_QUERY_FILE, read_only=True, data_only=True)
    ws = wb["Sell in Query"]
    headers = header_row(ws, 1)
    idx = {h: i for i, h in enumerate(headers)}
    sell_in = defaultdict(lambda: [0, 0, 0])
    exceptions = defaultdict(lambda: {"rows": 0, "qty": 0})
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        source_month = str(row["Cal. year / month"])
        if source_month not in SELL_IN_MONTHS:
            continue
        customer = as_int(row["Customer"])
        product = product_names.get(norm_name(row["Product Name (CloudViu)"]))
        if customer is None or not product:
            continue
        qty = row["PCS."] or 0
        qty = max(qty, 0)
        store_key = canonical.get(customer)
        if store_key is None:
            key = (customer, row.get("Customer Name"), product, "Customer code not in Final Master")
            exceptions[key]["rows"] += 1
            exceptions[key]["qty"] += qty
            continue
        sell_in[(store_key, product)][SELL_IN_MONTHS.index(source_month)] += qty
    wb.close()
    return sell_in, exceptions


def build_rows():
    stores, by_code, canonical = load_store_master()
    product_by_exact, product_by_norm = load_product_master()
    rsp_by_name = load_rsp()
    products, stock, raw_store_keys, stock_exceptions = load_raw_stock(stores, by_code, canonical, product_by_exact, product_by_norm, rsp_by_name)
    sell_in, sell_in_exceptions = load_sell_in(stores, by_code, canonical, products)

    rows = []
    for store_key in sorted(stores, key=lambda item: (item[0], item[1])):
        store = stores[store_key]
        store_code = store["_store_code"]
        store_name = store.get("Store Name (SCVD)") or store.get("Store Name (CloudViu)") or store.get("Store Name (Group)")
        store_group = store.get("Store Name (Group)") or store_name
        for product_name in sorted(products):
            product = products[product_name]
            current = stock.get((store_key, product_name), {}).get("current", 0)
            new = stock.get((store_key, product_name), {}).get("new", 0)
            total = current + new
            rsp = product["rsp"] or 0
            sell_values = sell_in.get((store_key, product_name), [0, 0, 0])
            sell_avg = sum(sell_values) / 3
            rows.append(
                [
                    store.get("Area"),
                    store.get("Sales Region (TT)"),
                    product["P_Group"],
                    product["Category"],
                    product["P_Code"],
                    product["Product SKU"],
                    product["Product Name"],
                    product["Size"],
                    product["Brand"],
                    product["Segment"],
                    product["Group"],
                    store_code,
                    store.get("Group Code"),
                    store.get("Channel"),
                    store.get("Type"),
                    store_name,
                    store.get("Province (EN)"),
                    REPORT_DATE,
                    store_code,
                    product["product_id"],
                    f"{store_code}{product['Product Name']}",
                    f"{product['Product SKU']}{store_group}",
                    rsp,
                    current,
                    new,
                    total,
                    total * rsp,
                    sell_values[0],
                    sell_values[1],
                    sell_values[2],
                    sell_avg,
                    sell_avg * rsp,
                    MONTH_LABEL,
                    YEAR,
                ]
            )
    return rows, {
        "raw_stores": len(raw_store_keys),
        "stores": len(stores),
        "products": len(products),
        "stock_exceptions": stock_exceptions,
        "sell_in_exceptions": sell_in_exceptions,
    }


def style(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["G"].width = 42
    ws.column_dimensions["P"].width = 34
    ws.column_dimensions["U"].width = 42
    ws.column_dimensions["V"].width = 42
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"


def write_exception_sheet(wb, title, headers, items):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for key, values in sorted(items.items(), key=lambda item: tuple(str(part) for part in item[0])):
        ws.append(list(key) + [values["rows"], values["qty"]])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="9C6500")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22


def write_workbook(rows, meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot"
    tt_rows = [row for row in rows if row[13] == "TT"]
    tt_stock_baht = sum(row[26] for row in tt_rows)
    tt_sell_baht = sum(row[31] for row in tt_rows)
    ws.append(["Channel", "TT", None])
    ws.append([None, None, None])
    ws.append(["Sum of Total Stock (Baht)", "Sum of  Total Sell in (Baht) ", "CVD"])
    ws.append([tt_stock_baht, tt_sell_baht, tt_stock_baht / tt_sell_baht * 30 if tt_sell_baht else 0])
    for cell in ws[3]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 12

    ws = wb.create_sheet("Raw Data")
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    style(ws)

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    metrics = [
        ("Month", MONTH_LABEL),
        ("Rows", len(rows)),
        ("Store instances", meta["stores"]),
        ("Stores with raw stock", meta["raw_stores"]),
        ("SKUs from raw May", meta["products"]),
        ("Expected crossjoin", meta["stores"] * meta["products"]),
        ("Total stock qty", sum(row[25] for row in rows)),
        ("Total stock baht", sum(row[26] for row in rows)),
        ("Sell-in qty", sum(row[30] for row in rows)),
        ("Sell-in baht", sum(row[31] for row in rows)),
        ("Unmapped stock rows", sum(v["rows"] for v in meta["stock_exceptions"].values())),
        ("Unmapped stock qty", sum(v["qty"] for v in meta["stock_exceptions"].values())),
        ("Unmapped sell-in rows", sum(v["rows"] for v in meta["sell_in_exceptions"].values())),
        ("Unmapped sell-in qty", sum(v["qty"] for v in meta["sell_in_exceptions"].values())),
    ]
    for metric in metrics:
        summary.append(metric)
    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 18
    write_exception_sheet(
        wb,
        "Stock Mapping Exceptions",
        ["Store Code", "Outlet ID", "Outlet Name", "Product Name", "Reason", "Source Rows", "Qty"],
        meta["stock_exceptions"],
    )
    write_exception_sheet(
        wb,
        "Sell-in Mapping Exceptions",
        ["Customer Code", "Customer Name", "Product Name", "Reason", "Source Rows", "Qty"],
        meta["sell_in_exceptions"],
    )
    wb.save(MONTHLY_OUTPUT)


def main():
    rows, meta = build_rows()
    print(
        {
            **meta,
            "rows": len(rows),
            "expected_crossjoin": meta["stores"] * meta["products"],
            "stock_qty": sum(row[25] for row in rows),
            "stock_baht": sum(row[26] for row in rows),
            "sell_in_qty": sum(row[30] for row in rows),
            "sell_in_baht": sum(row[31] for row in rows),
            "unmapped_stock_rows": sum(v["rows"] for v in meta["stock_exceptions"].values()),
            "unmapped_stock_qty": sum(v["qty"] for v in meta["stock_exceptions"].values()),
            "unmapped_sell_in_rows": sum(v["rows"] for v in meta["sell_in_exceptions"].values()),
            "unmapped_sell_in_qty": sum(v["qty"] for v in meta["sell_in_exceptions"].values()),
        }
    )
    write_workbook(rows, meta)
    print(f"OUTPUT={MONTHLY_OUTPUT}")


if __name__ == "__main__":
    main()
