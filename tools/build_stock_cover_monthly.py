from collections import defaultdict
from datetime import datetime
from pathlib import Path
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\Work\Nestle")
REPORT_DIR = BASE / r"Report\(5th) Stock Cover Day"
MASTER_FILE = BASE / r"Master\Final Master.xlsx"
RSP_FILE = BASE / r"Master\RSP on month.xlsx"
STAGING_FILE = REPORT_DIR / "Stock Cover Day.xlsx"

MONTH_LABEL = "05_May-26"
SELL_IN_MONTH = "05.2026"
REPORT_DATE = datetime(2026, 6, 1)
YEAR = 2026

MONTHLY_OUTPUT = REPORT_DIR / "Stock Monthly - May 2026.xlsx"
REPORT_OUTPUT = REPORT_DIR / "Stock Cover Day Report May 2026 - V.3.xlsx"

REPORT_HEADERS = [
    "Area",
    "New Region",
    "P_Group",
    "Category",
    "P_Code",
    "Product SKU",
    "Product Name",
    "Size",
    "Brand",
    "Segment",
    "Group",
    "Nestlé Code",
    "Nestlé Group Code",
    "Channel",
    "Type",
    "Store Name",
    "Province",
    "Date",
    "outlet_id",
    "product_id",
    "Concat",
    "Concat2",
    " RSP LTP = TT (-1) ",
    "#Stock คงเหลือ โฉมปัจจุบัน",
    "#Stock คงเหลือ โฉมใหม่",
    "#Total Stock",
    "Total Stock (Baht)",
    "Sell in = TT (1)",
    "Sell in = TT (2)",
    "Sell in = TT (3)",
    "Sell in (1+2+3)/3",
    " Total Sell in (Baht) ",
    "Month",
    "Year",
]


def as_int(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def number_parts(value):
    return [int(item) for item in re.findall(r"\d+", str(value or ""))]


def header_from_row(ws, row_num):
    return [cell.value for cell in next(ws.iter_rows(min_row=row_num, max_row=row_num))]


def row_dicts(ws, header_row, min_row=None):
    headers = header_from_row(ws, header_row)
    min_row = min_row or header_row + 1
    for values in ws.iter_rows(min_row=min_row, values_only=True):
        if all(value is None for value in values):
            continue
        yield {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}


def load_store_master():
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws = wb["Store Master"]

    code_to_latest = {}
    latest_rows = {}
    for row in row_dicts(ws, 15, 16):
        store_code = as_int(row.get("Store Code"))
        if store_code is None:
            continue
        candidates = number_parts(row.get("Group Code")) or [store_code]
        latest_code = max(candidates)
        code_to_latest[store_code] = latest_code

        existing = latest_rows.get(latest_code)
        existing_code = as_int(existing.get("Store Code")) if existing else None
        if existing is None or (store_code == latest_code and existing_code != latest_code):
            latest_rows[latest_code] = row

    wb.close()
    return code_to_latest, latest_rows


def normalize_product_name(name):
    return str(name or "").replace(" (โฉมใหม่)", "").strip()


def load_product_master():
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws = wb["Product Master"]
    by_name = {}
    for row in row_dicts(ws, 14, 15):
        name = normalize_product_name(row.get("Product Name (โฉมใหม่)") or row.get("Product Name") or row.get("name"))
        if not name:
            continue
        if name not in by_name:
            by_name[name] = row
    wb.close()
    return by_name


def load_rsp():
    wb = openpyxl.load_workbook(RSP_FILE, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = header_from_row(ws, 1)
    idx = {h: i for i, h in enumerate(headers)}
    rsp = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Month"]] == MONTH_LABEL:
            rsp[normalize_product_name(row[idx["Product Name"]])] = row[idx["RSP"]] or 0
    wb.close()
    return rsp


def load_stock_staging(code_to_latest, product_master, rsp_by_name):
    wb = openpyxl.load_workbook(STAGING_FILE, read_only=True, data_only=True)
    ws = wb["Stock"]
    headers = header_from_row(ws, 1)
    idx = {h: i for i, h in enumerate(headers)}

    products = {}
    stock = defaultdict(lambda: {"current": 0, "new": 0})

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_code = as_int(row[idx["Nestlé Code"]])
        if raw_code is None:
            continue
        latest_code = code_to_latest.get(raw_code, raw_code)
        product_name = normalize_product_name(row[idx["item_product_name"]])
        if not product_name:
            continue

        master = product_master.get(product_name, {})
        product = products.setdefault(
            product_name,
            {
                "P_Group": master.get("P_group") or row[idx["P_Group"]],
                "Category": master.get("Category") or row[idx["Category"]],
                "P_Code": master.get("P_code") or row[idx["P_Code"]],
                "Product SKU": master.get("Attribute Group") or row[idx["item_label"]],
                "Product Name": product_name,
                "Size": master.get("Size") or row[idx["Size"]],
                "Brand": master.get("Brand") or row[idx["item_label"]],
                "Segment": master.get("Group") or row[idx["Category"]],
                "Group": master.get("Type") or row[idx["Group"]],
                "product_id": master.get("sku"),
                "material": as_int(master.get("Nestle product Code")),
                "rsp": rsp_by_name.get(product_name) or master.get("marketPrice") or row[idx["RSP LTP = TT (-1)"]] or 0,
            },
        )
        if not product.get("product_id"):
            product["product_id"] = master.get("sku")
        if not product.get("material"):
            product["material"] = as_int(master.get("Nestle product Code"))

        stock[(latest_code, product_name)]["current"] += row[idx["โฉมปัจจุบัน"]] or 0
        stock[(latest_code, product_name)]["new"] += row[idx["โฉมใหม่"]] or 0

    wb.close()
    return products, stock


def load_sell_in(code_to_latest, products):
    product_names = {normalize_product_name(name): name for name in products}

    wb = openpyxl.load_workbook(STAGING_FILE, read_only=True, data_only=True)
    ws = wb["Sell in"]
    headers = header_from_row(ws, 1)
    idx = {h: i for i, h in enumerate(headers)}
    sell_in = defaultdict(float)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[idx["Cal. year / month"]]) != SELL_IN_MONTH:
            continue
        customer = as_int(row[idx["Customer"]])
        if customer is None:
            continue
        latest_code = code_to_latest.get(customer, customer)
        product_name = product_names.get(normalize_product_name(row[idx["Product Name (CloudViu)"]]))
        if product_name:
            sell_in[(latest_code, product_name)] += row[idx["PCS."]] or 0

    wb.close()
    return sell_in


def build_rows():
    code_to_latest, stores = load_store_master()
    product_master = load_product_master()
    rsp_by_name = load_rsp()
    products, stock = load_stock_staging(code_to_latest, product_master, rsp_by_name)
    sell_in = load_sell_in(code_to_latest, products)

    rows = []
    for store_code in sorted(stores):
        store = stores[store_code]
        for product_name in sorted(products):
            product = products[product_name]
            stock_values = stock.get((store_code, product_name), {"current": 0, "new": 0})
            current_stock = stock_values["current"]
            new_stock = stock_values["new"]
            total_stock = current_stock + new_stock
            rsp = product["rsp"] or 0
            stock_baht = total_stock * rsp
            sell_qty = sell_in.get((store_code, product_name), 0)
            sell_in_baht = sell_qty * rsp
            store_name = store.get("Store Name (SCVD)") or store.get("Store Name (CloudViu)") or store.get("Store Name (Group)")
            store_group = store.get("Store Name (Group)") or store_name
            rows.append(
                [
                    store.get("Area"),
                    store.get("Sales Region (TT)"),
                    product["P_Group"],
                    product["Category"],
                    product["P_Code"],
                    product["Product SKU"],
                    product["Product Name"],
                    product["Size"],
                    product["Brand"],
                    product["Segment"],
                    product["Group"],
                    store_code,
                    store.get("Group Code"),
                    store.get("Channel"),
                    store.get("Type"),
                    store_name,
                    store.get("Province (EN)"),
                    REPORT_DATE,
                    store_code,
                    product["product_id"],
                    f"{store_code}{product['Product Name']}",
                    f"{product['Product SKU']}{store_group}",
                    rsp,
                    current_stock,
                    new_stock,
                    total_stock,
                    stock_baht,
                    0,
                    0,
                    sell_qty,
                    sell_qty,
                    sell_in_baht,
                    MONTH_LABEL,
                    YEAR,
                ]
            )
    return rows


def style_raw_sheet(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        1: 10,
        2: 16,
        3: 24,
        4: 22,
        6: 28,
        7: 42,
        16: 34,
        21: 42,
        22: 42,
    }
    for col in range(1, len(REPORT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 14)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(REPORT_HEADERS))}{ws.max_row}"


def write_monthly(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"
    ws.append(REPORT_HEADERS)
    for row in rows:
        ws.append(row)
    style_raw_sheet(ws)

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary_rows = [
        ("Month", MONTH_LABEL),
        ("Rows", len(rows)),
        ("Stores", len({row[11] for row in rows})),
        ("Products", len({row[6] for row in rows})),
        ("Total Stock", sum(row[25] for row in rows)),
        ("Total Stock (Baht)", sum(row[26] for row in rows)),
        ("Sell-in Quantity", sum(row[30] for row in rows)),
        ("Total Sell-in (Baht)", sum(row[31] for row in rows)),
    ]
    for item in summary_rows:
        summary.append(item)
    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 18
    wb.save(MONTHLY_OUTPUT)


def write_report(rows):
    wb = Workbook()

    def add_table(sheet_name, headers, data_rows):
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for item in data_rows:
            ws.append(item)
        style_raw_sheet(ws)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = min(max(len(str(headers[col - 1])) + 2, 12), 32)
        return ws

    wb.remove(wb.active)

    channel = defaultdict(lambda: [0, set(), set(), 0, 0, 0, 0])
    sub_brand = defaultdict(lambda: [0, set(), set(), 0, 0, 0, 0])
    store = defaultdict(lambda: [None, None, None, None, set(), 0, 0, 0, 0])
    store_sku = defaultdict(lambda: [None, None, None, None, None, 0, 0, 0, 0])

    for row in rows:
        ch = row[13]
        channel[ch][0] += 1
        channel[ch][1].add(row[11])
        channel[ch][2].add(row[6])
        channel[ch][3] += row[25]
        channel[ch][4] += row[26]
        channel[ch][5] += row[30]
        channel[ch][6] += row[31]

        sb_key = (row[2], row[3])
        sub_brand[sb_key][0] += 1
        sub_brand[sb_key][1].add(row[11])
        sub_brand[sb_key][2].add(row[6])
        sub_brand[sb_key][3] += row[25]
        sub_brand[sb_key][4] += row[26]
        sub_brand[sb_key][5] += row[30]
        sub_brand[sb_key][6] += row[31]

        st_key = row[11]
        store[st_key][0] = row[0]
        store[st_key][1] = row[1]
        store[st_key][2] = row[15]
        store[st_key][3] = row[13]
        store[st_key][4].add(row[6])
        store[st_key][5] += row[25]
        store[st_key][6] += row[26]
        store[st_key][7] += row[30]
        store[st_key][8] += row[31]

        ss_key = (row[11], row[6])
        store_sku[ss_key][0] = row[11]
        store_sku[ss_key][1] = row[15]
        store_sku[ss_key][2] = row[6]
        store_sku[ss_key][3] = row[2]
        store_sku[ss_key][4] = row[13]
        store_sku[ss_key][5] += row[25]
        store_sku[ss_key][6] += row[26]
        store_sku[ss_key][7] += row[30]
        store_sku[ss_key][8] += row[31]

    def cvd(stock_baht, sell_baht):
        return (stock_baht / sell_baht * 30) if sell_baht else 0

    add_table(
        "Summary By Channel",
        ["Channel", "Rows", "Stores", "SKUs", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"],
        [
            [k, v[0], len(v[1]), len(v[2]), v[3], v[4], v[5], v[6], cvd(v[4], v[6])]
            for k, v in sorted(channel.items())
        ],
    )
    add_table(
        "Summary By Sub Brand",
        ["P_Group", "Category", "Rows", "Stores", "SKUs", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"],
        [
            [k[0], k[1], v[0], len(v[1]), len(v[2]), v[3], v[4], v[5], v[6], cvd(v[4], v[6])]
            for k, v in sorted(sub_brand.items())
        ],
    )
    add_table(
        "Stock Data",
        ["Area", "New Region", "Nestle Code", "Channel", "Store Name", "SKUs", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"],
        [
            [v[0], v[1], k, v[3], v[2], len(v[4]), v[5], v[6], v[7], v[8], cvd(v[6], v[8])]
            for k, v in sorted(store.items())
        ],
    )
    add_table(
        "Raw Data",
        REPORT_HEADERS,
        rows,
    )
    add_table(
        "Pivot by Store",
        ["Nestle Code", "Store Name", "Channel", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"],
        [
            [k, v[2], v[3], v[5], v[6], v[7], v[8], cvd(v[6], v[8])]
            for k, v in sorted(store.items())
        ],
    )
    add_table(
        "Store by SKU",
        ["Nestle Code", "Store Name", "Product Name", "P_Group", "Channel", "Total Stock", "Total Stock (Baht)", "Sell-in Qty", "Total Sell-in (Baht)", "CVD"],
        [
            [v[0], v[1], v[2], v[3], v[4], v[5], v[6], v[7], v[8], cvd(v[6], v[8])]
            for _k, v in sorted(store_sku.items())
        ],
    )
    wb.save(REPORT_OUTPUT)
    wb.close()


def validate(rows):
    return {
        "rows": len(rows),
        "stores": len({row[11] for row in rows}),
        "products": len({row[6] for row in rows}),
        "expected_crossjoin": len({row[11] for row in rows}) * len({row[6] for row in rows}),
        "stock_qty": sum(row[25] for row in rows),
        "stock_baht": sum(row[26] for row in rows),
        "sell_in_qty": sum(row[30] for row in rows),
        "sell_in_baht": sum(row[31] for row in rows),
    }


def main():
    rows = build_rows()
    print(validate(rows))
    write_monthly(rows)
    write_report(rows)
    print(f"MONTHLY_OUTPUT={MONTHLY_OUTPUT}")
    print(f"REPORT_OUTPUT={REPORT_OUTPUT}")


if __name__ == "__main__":
    main()
