import collections
import json

import openpyxl


RAW_FILE = r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Raw Data for Stock Cover Day Report May 2026.xlsx"
SELLIN_FILE = r"C:\Work\Nestle\Target & Incentive\Sell in Query.xlsx"
MASTER_FILE = r"C:\Work\Nestle\Master\Final Master.xlsx"


def headers(ws, row=1):
    h = [c.value for c in next(ws.iter_rows(min_row=row, max_row=row))]
    return h, {v: i for i, v in enumerate(h)}


def n(v):
    return v if isinstance(v, (int, float)) else 0


def norm(v):
    return str(v or "").strip()


def raw_cnt_uht_summary():
    wb = openpyxl.load_workbook(RAW_FILE, read_only=True, data_only=True)
    ws = wb["Raw Data"]
    h, idx = headers(ws)
    by_type_month = collections.defaultdict(lambda: {"stock_baht": 0, "sell_baht": 0, "sell1": 0, "sell2": 0, "sell3": 0, "rows": 0, "products": set()})
    products = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["P_Group"]] != "CNT UHT":
            continue
        key = (row[idx["Type"]], row[idx["Month"]])
        b = by_type_month[key]
        b["rows"] += 1
        b["products"].add(row[idx["Product Name"]])
        products.add(row[idx["Product Name"]])
        b["stock_baht"] += n(row[idx["Total Stock (Baht)"]])
        b["sell_baht"] += n(row[idx[" Total Sell in (Baht) "]])
        b["sell1"] += n(row[idx["Sell in = TT (1)"]])
        b["sell2"] += n(row[idx["Sell in = TT (2)"]])
        b["sell3"] += n(row[idx["Sell in = TT (3)"]])
    wb.close()
    return {
        "products": sorted(products),
        "by_type_month": [
            [typ, month, data["rows"], len(data["products"]), data["stock_baht"], data["sell1"], data["sell2"], data["sell3"], data["sell_baht"]]
            for (typ, month), data in sorted(by_type_month.items(), key=lambda item: (str(item[0][0]), str(item[0][1])))
        ],
    }


def product_master_cnt_names():
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws = wb["Product Master"]
    h, idx = headers(ws, 14)
    names = []
    for row in ws.iter_rows(min_row=15, values_only=True):
        p_group = row[idx.get("P_group")]
        if p_group != "CNT UHT":
            continue
        names.append(
            {
                "name": row[idx.get("name")],
                "Product Name": row[idx.get("Product Name")],
                "Product Name New": row[idx.get("Product Name (โฉมใหม่)")],
                "Attribute Group": row[idx.get("Attribute Group")],
                "P_code": row[idx.get("P_code")],
            }
        )
    wb.close()
    return names


def sellin_cnt_uht_raw(raw_products):
    product_set = {norm(p) for p in raw_products}
    wb = openpyxl.load_workbook(SELLIN_FILE, read_only=True, data_only=True)
    ws = wb["Sell in Query"]
    h, idx = headers(ws)
    by_month_pgroup = collections.defaultdict(lambda: {"pcs": 0, "nps": 0, "rows": 0, "products": set()})
    by_month_cloud = collections.defaultdict(lambda: {"pcs": 0, "nps": 0, "rows": 0})
    cnt_like_rows = collections.defaultdict(lambda: {"pcs": 0, "nps": 0, "rows": 0, "cloud_products": set(), "sap_products": set(), "p_groups": set()})
    raw_match_by_month = collections.defaultdict(lambda: {"pcs": 0, "nps": 0, "rows": 0, "products": set()})
    months = {f"{m:02d}.2025" for m in range(1, 13)} | {f"{m:02d}.2026" for m in range(1, 6)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        month = str(row[idx["Cal. year / month"]] or "")
        if month not in months:
            continue
        cloud = norm(row[idx["Product Name (CloudViu)"]])
        p_group = norm(row[idx["P_Group"]])
        group_scvd = norm(row[idx["Group for SCVD"]])
        category = norm(row[idx["Category"]])
        sap = norm(row[idx["Product Name"]])
        pcs = max(n(row[idx["PCS."]]), 0)
        nps = n(row[idx["NPS - THB"]])
        if p_group == "CNT UHT":
            b = by_month_pgroup[month]
            b["pcs"] += pcs
            b["nps"] += nps
            b["rows"] += 1
            b["products"].add(cloud)
        if cloud in product_set:
            b = raw_match_by_month[month]
            b["pcs"] += pcs
            b["nps"] += nps
            b["rows"] += 1
            b["products"].add(cloud)
        if "Carnation 3 Smartgo UHT" in cloud or "Carnation 3 Smartgo UHT" in sap or p_group == "CNT UHT" or "CNT UHT" in group_scvd:
            b = cnt_like_rows[month]
            b["pcs"] += pcs
            b["nps"] += nps
            b["rows"] += 1
            b["cloud_products"].add(cloud)
            b["sap_products"].add(sap)
            b["p_groups"].add(p_group)

    wb.close()
    return {
        "by_month_pgroup_cnt_uht": {
            m: {"pcs": v["pcs"], "nps": v["nps"], "rows": v["rows"], "products": sorted(v["products"])}
            for m, v in sorted(by_month_pgroup.items())
        },
        "raw_product_name_matches_by_month": {
            m: {"pcs": v["pcs"], "nps": v["nps"], "rows": v["rows"], "products": sorted(v["products"])}
            for m, v in sorted(raw_match_by_month.items())
        },
        "cnt_like_by_month": {
            m: {
                "pcs": v["pcs"],
                "nps": v["nps"],
                "rows": v["rows"],
                "cloud_products": sorted(v["cloud_products"]),
                "p_groups": sorted(v["p_groups"]),
                "sap_sample": sorted(v["sap_products"])[:10],
            }
            for m, v in sorted(cnt_like_rows.items())
        },
    }


def main():
    raw = raw_cnt_uht_summary()
    result = {
        "raw_cnt_uht_products": raw["products"],
        "raw_by_type_month": raw["by_type_month"],
        "product_master_cnt_uht": product_master_cnt_names(),
        "sellin_query": sellin_cnt_uht_raw(raw["products"]),
    }
    print(json.dumps(result, ensure_ascii=False, default=str, indent=2))


if __name__ == "__main__":
    main()
