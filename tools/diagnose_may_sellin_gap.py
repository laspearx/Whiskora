import collections
import json

import openpyxl


MONTHLY = r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Monthly - May 2026.xlsx"
V2 = r"C:\Work\Nestle\Report\(5th) Stock Cover Day\Stock Cover Day Report May 2026 - V.2.xlsx"

CODE_HEADERS = ["Nestlé Code", "Nestlรฉ Code", "Nestlเธฃเธ\x89 Code"]
SELL_QTY_HEADERS = ["Sell in (1+2+3)/3", "(Last 3 Months)/3", "(Aug'25-Oct'25)/3"]


def header_map(ws):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return headers, {header: i for i, header in enumerate(headers)}


def num(row, idx, col):
    if col not in idx:
        return 0
    value = row[idx[col]]
    return value if isinstance(value, (int, float)) else 0


def code_col(idx):
    for col in CODE_HEADERS:
        if col in idx:
            return col
    raise KeyError("No store code column found")


def load_raw(path, label):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Raw Data"]
    headers, idx = header_map(ws)
    ccol = code_col(idx)

    by_product = collections.defaultdict(lambda: {"sell_baht": 0, "sell_qty": 0, "stock_baht": 0, "rows": 0})
    by_store = collections.defaultdict(lambda: {"sell_baht": 0, "sell_qty": 0, "stock_baht": 0, "rows": 0, "type": None})
    by_store_product = collections.defaultdict(lambda: {"sell_baht": 0, "sell_qty": 0, "rows": 0})
    by_code_product_store = collections.defaultdict(list)
    total = {"sell_baht": 0, "sell_qty": 0, "stock_baht": 0, "rows": 0}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if "Month" in idx and label == "v2" and row[idx["Month"]] != "05_May-26":
            continue
        if row[idx["Channel"]] != "TT":
            continue
        code = row[idx[ccol]]
        store = row[idx["Store Name"]]
        product = row[idx["Product Name"]]
        typ = row[idx["Type"]]
        sell_qty = sum(num(row, idx, col) for col in SELL_QTY_HEADERS)
        sell_baht = num(row, idx, " Total Sell in (Baht) ")
        stock_baht = num(row, idx, "Total Stock (Baht)")

        total["rows"] += 1
        total["sell_qty"] += sell_qty
        total["sell_baht"] += sell_baht
        total["stock_baht"] += stock_baht

        by_product[product]["rows"] += 1
        by_product[product]["sell_qty"] += sell_qty
        by_product[product]["sell_baht"] += sell_baht
        by_product[product]["stock_baht"] += stock_baht

        by_store[(code, store)]["rows"] += 1
        by_store[(code, store)]["type"] = typ
        by_store[(code, store)]["sell_qty"] += sell_qty
        by_store[(code, store)]["sell_baht"] += sell_baht
        by_store[(code, store)]["stock_baht"] += stock_baht

        by_store_product[(code, store, product)]["rows"] += 1
        by_store_product[(code, store, product)]["sell_qty"] += sell_qty
        by_store_product[(code, store, product)]["sell_baht"] += sell_baht

    for (code, store, product), data in by_store_product.items():
        if data["sell_baht"]:
            by_code_product_store[(code, product)].append((store, data["sell_qty"], data["sell_baht"]))

    wb.close()
    return {
        "total": total,
        "by_product": by_product,
        "by_store": by_store,
        "by_store_product": by_store_product,
        "by_code_product_store": by_code_product_store,
    }


def top_product_gap(monthly, v2):
    rows = []
    for product in set(monthly) | set(v2):
        m = monthly.get(product, {})
        o = v2.get(product, {})
        rows.append([
            product,
            o.get("sell_baht", 0),
            m.get("sell_baht", 0),
            m.get("sell_baht", 0) - o.get("sell_baht", 0),
            o.get("sell_qty", 0),
            m.get("sell_qty", 0),
            m.get("sell_qty", 0) - o.get("sell_qty", 0),
        ])
    return sorted(rows, key=lambda row: abs(row[3]), reverse=True)


def top_store_gap(monthly, v2):
    rows = []
    for key in set(monthly) | set(v2):
        m = monthly.get(key, {})
        o = v2.get(key, {})
        rows.append([
            key[0],
            key[1],
            o.get("type"),
            o.get("sell_baht", 0),
            m.get("sell_baht", 0),
            m.get("sell_baht", 0) - o.get("sell_baht", 0),
            o.get("sell_qty", 0),
            m.get("sell_qty", 0),
            m.get("sell_qty", 0) - o.get("sell_qty", 0),
        ])
    return sorted(rows, key=lambda row: abs(row[5]), reverse=True)


def duplicate_sellin_by_code(data):
    duplicates = []
    possible_over = 0
    for (code, product), store_rows in data["by_code_product_store"].items():
        stores = [row for row in store_rows if row[2]]
        if len(stores) <= 1:
            continue
        total = sum(row[2] for row in stores)
        max_one = max(row[2] for row in stores)
        possible_over += total - max_one
        duplicates.append([code, product, len(stores), total, max_one, total - max_one, stores])
    return possible_over, sorted(duplicates, key=lambda row: row[5], reverse=True)


def exception_summary():
    wb = openpyxl.load_workbook(MONTHLY, read_only=True, data_only=True)
    out = {}
    for sheet in ["Sell-in Mapping Exceptions"]:
        ws = wb[sheet]
        headers, idx = header_map(ws)
        by_customer = collections.defaultdict(lambda: {"qty": 0, "rows": 0, "name": None})
        by_product = collections.defaultdict(float)
        for row in ws.iter_rows(min_row=2, values_only=True):
            customer = row[idx["Customer Code"]]
            product = row[idx["Product Name"]]
            qty = row[-1] or 0
            rows = row[-2] or 0
            by_customer[customer]["name"] = row[idx["Customer Name"]]
            by_customer[customer]["qty"] += qty
            by_customer[customer]["rows"] += rows
            by_product[product] += qty
        out["top_customer_qty"] = sorted(
            [[code, data["name"], data["rows"], data["qty"]] for code, data in by_customer.items()],
            key=lambda row: row[3],
            reverse=True,
        )[:20]
        out["top_product_qty"] = sorted([[product, qty] for product, qty in by_product.items()], key=lambda row: row[1], reverse=True)[:20]
    wb.close()
    return out


def main():
    monthly = load_raw(MONTHLY, "monthly")
    v2 = load_raw(V2, "v2")
    product_gaps = top_product_gap(monthly["by_product"], v2["by_product"])
    store_gaps = top_store_gap(monthly["by_store"], v2["by_store"])
    v2_dup_over, v2_dups = duplicate_sellin_by_code(v2)
    monthly_dup_over, monthly_dups = duplicate_sellin_by_code(monthly)

    v2_products = set(v2["by_product"])
    monthly_products = set(monthly["by_product"])
    v2_only = [row for row in product_gaps if row[0] in v2_products - monthly_products]
    monthly_only = [row for row in product_gaps if row[0] in monthly_products - v2_products]

    result = {
        "totals": {"monthly": monthly["total"], "v2": v2["total"], "diff_monthly_minus_v2": {k: monthly["total"][k] - v2["total"][k] for k in monthly["total"]}},
        "top_product_gap": product_gaps[:30],
        "v2_only_products": v2_only[:30],
        "monthly_only_products": monthly_only[:30],
        "v2_only_product_sell_baht_total": sum(row[1] for row in v2_only),
        "monthly_only_product_sell_baht_total": sum(row[2] for row in monthly_only),
        "top_store_gap": store_gaps[:30],
        "possible_v2_duplicate_sell_baht_by_same_code_product": v2_dup_over,
        "possible_monthly_duplicate_sell_baht_by_same_code_product": monthly_dup_over,
        "top_v2_duplicate_code_product": v2_dups[:20],
        "top_monthly_duplicate_code_product": monthly_dups[:20],
        "monthly_sellin_exceptions": exception_summary(),
    }
    print(json.dumps(result, ensure_ascii=False, default=str, indent=2))


if __name__ == "__main__":
    main()
